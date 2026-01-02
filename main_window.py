import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
from pathlib import Path
import os
import json
import logging
import time
from typing import Dict, List, Optional
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed, wait, FIRST_COMPLETED
import queue
from core.tester import PageLoadTester
from core.result import TestResult, TestSession, AverageResult
from core.blacklist_manager import BlacklistManager
from utils.logger import setup_logger
from utils.file_utils import save_json_file, load_json_file, save_excel_file, save_average_results_to_excel, get_app_base_dir
from utils.config_manager import ConfigManager
from ui.config_dialog import ConfigDialog
from ui.unified_test_dialog import show_unified_test_dialog
from ui.schedule_manager import show_schedule_manager
import threading

class TextHandler(logging.Handler):
    """自定义日志处理器，将日志输出到文本框（线程安全）"""

    def __init__(self, text_widget, root=None):
        super().__init__()
        self.text_widget = text_widget
        self.root = root  # 保存root引用，用于线程安全调用
        self.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
        self._log_queue = queue.Queue()  # 用于线程安全的日志队列
        self._setup_queue_processor()

    def _setup_queue_processor(self):
        """设置队列处理器，在主线程中定期处理日志消息"""
        def process_queue():
            try:
                while True:
                    msg = self._log_queue.get_nowait()
                    try:
                        self.text_widget.config(state=tk.NORMAL)
                        self.text_widget.insert(tk.END, msg + '\n')
                        self.text_widget.see(tk.END)  # 自动滚动到底部
                        self.text_widget.config(state=tk.DISABLED)
                    except (RuntimeError, tk.TclError):
                        # 如果UI已经关闭或不可用，忽略错误
                        pass
            except queue.Empty:
                pass
            finally:
                # 继续处理队列（每100ms检查一次）
                if self.root:
                    try:
                        self.root.after(100, process_queue)
                    except (RuntimeError, tk.TclError):
                        pass

        # 启动队列处理器
        if self.root:
            try:
                self.root.after(100, process_queue)
            except (RuntimeError, tk.TclError):
                pass

    def emit(self, record):
        msg = self.format(record)

        # 检查是否在主线程中
        if threading.current_thread() is threading.main_thread():
            # 在主线程中，直接更新UI
            try:
                self.text_widget.config(state=tk.NORMAL)
                self.text_widget.insert(tk.END, msg + '\n')
                self.text_widget.see(tk.END)  # 自动滚动到底部
                self.text_widget.config(state=tk.DISABLED)
            except (RuntimeError, tk.TclError):
                # 如果UI已经关闭或不可用，忽略错误
                pass
        else:
            # 在后台线程中，将消息放入队列
            try:
                self._log_queue.put_nowait(msg)
            except queue.Full:
                # 如果队列满了，忽略这条日志（避免阻塞）
                pass


class MainWindow:
    def __init__(self, root: tk.Tk, logger: logging.Logger):
        """初始化主窗口"""
        self.root = root
        self.root.title("网页性能测试工具")
        # 设置默认窗口大小，但允许调整（更紧凑）
        self.root.geometry("1100x700")
        # 设置最小尺寸，适配当前精简布局
        self.root.minsize(900, 550)

        # 依赖注入：日志和测试器
        self.logger = logger
        self.config_manager = ConfigManager()
        self.tester = PageLoadTester(config_manager=self.config_manager)
        
        # 黑名单管理器（使用应用基础目录下的config/blacklist.json）
        blacklist_file = get_app_base_dir() / "config" / "blacklist.json"
        self.blacklist_manager = BlacklistManager(blacklist_file=blacklist_file)
        self.tester.blacklist_manager = self.blacklist_manager
        # 记录黑名单加载状态
        blocked_count = len(self.blacklist_manager.get_blocked_domains())
        if blocked_count > 0:
            self.logger.info(f"已加载黑名单: {blocked_count} 个域名")
        else:
            self.logger.info(f"黑名单文件不存在或为空: {blacklist_file}")
        
        # 定时任务调度器
        from utils.scheduler import TaskScheduler
        self.scheduler = TaskScheduler(logger=self.logger)
        # 恢复保存的任务
        self.scheduler.restore_tasks(lambda config: lambda c=config: self._execute_scheduled_test(c))
        self.scheduler.start()

        # 结果存储 - 更新类型
        self.current_session_name = ""
        self.current_session_results: List[TestResult] = []
        self.all_sessions: Dict[str, List[TestResult]] = {}
        self.session_objects: Dict[str, TestSession] = {}
        
        # 文件夹和文件选择相关
        self.selected_folder_path = None  # 当前选择的文件夹路径
        self.folder_files_map = {}  # 存储每个文件夹下的文件列表

        # 视图状态管理
        self.current_view = "detail"
        self.current_average_results: List[AverageResult] = []

        # 测试进度相关变量
        self.is_testing = False
        self.test_stopped = False  # 停止测试标志
        self.current_test_config = None
        self.current_test_session = None  # 当前测试会话（用于停止时保存）
        self.current_round = 0
        self.current_url_index = 0
        self.total_rounds = 0
        self.total_urls = 0
        self.test_threads: List[threading.Thread] = []
        self.testing_lock = threading.Lock()  # 保护is_testing状态的锁
        self.test_completion_events: Dict[str, threading.Event] = {}  # 用于按次数执行模式的测试完成事件

        # 创建UI
        self.create_widgets()

        # 设置自定义日志处理器
        self.setup_logging_handler()

        # 状态栏
        self.create_status_bar()

        # 加载历史结果（自动加载模式，不弹出对话框）
        # 不再自动加载本地会话
        # self.load_previous_results(auto_load=True)
        self.last_test_config = None
        
        # 绑定窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _start_managed_thread(self, target, name: str = "test-thread") -> threading.Thread:
        """启动受管的后台线程，确保退出时能正常等待"""
        # 清理已结束的线程引用
        self.test_threads = [t for t in self.test_threads if t.is_alive()]
        thread = threading.Thread(target=target, name=name, daemon=False)
        thread.start()
        self.test_threads.append(thread)
        return thread

    def _wait_for_test_threads(self, timeout: float = 5.0) -> None:
        """等待当前受管线程结束，避免异常退出"""
        for thread in list(self.test_threads):
            if thread.is_alive():
                thread.join(timeout=timeout)
        self.test_threads = [t for t in self.test_threads if t.is_alive()]

    def _safe_after(self, delay_ms: int, callback):
        """在窗口存在时才调度回调，防止主循环已退出"""
        try:
            if self.root and self.root.winfo_exists():
                self.root.after(delay_ms, callback)
        except tk.TclError:
            # 主循环已退出，忽略
            pass

    def setup_logging_handler(self):
        """设置自定义日志处理器"""
        # 保留文件处理器，只移除控制台处理器和文本框处理器（避免重复）
        handlers_to_remove = []
        for handler in self.logger.handlers[:]:
            # 保留文件处理器（FileHandler 或 DualFileHandler）
            if isinstance(handler, logging.FileHandler):
                continue
            # 移除控制台处理器和文本框处理器
            if isinstance(handler, (logging.StreamHandler, TextHandler)):
                handlers_to_remove.append(handler)
        
        for handler in handlers_to_remove:
            self.logger.removeHandler(handler)

        # 添加控制台处理器
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(
            logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
        self.logger.addHandler(console_handler)

        # 添加文本框处理器（传入root引用以确保线程安全）
        text_handler = TextHandler(self.log_text, root=self.root)
        text_handler.setLevel(logging.INFO)
        self.logger.addHandler(text_handler)

        self.logger.setLevel(logging.INFO)

    def create_status_bar(self):
        """创建状态栏"""
        status_frame = ttk.Frame(self.root)
        status_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=2)

        # 测试进度显示
        self.progress_var = tk.StringVar(value="")
        progress_label = ttk.Label(status_frame, textvariable=self.progress_var, font=("Arial", 9))
        progress_label.pack(side=tk.LEFT)

        # 状态信息显示
        self.status_var = tk.StringVar(value="就绪")
        status_label = ttk.Label(
            status_frame,
            textvariable=self.status_var,
            relief=tk.SUNKEN,
            anchor=tk.W,
            font=("Arial", 9)
        )
        status_label.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))

    def create_widgets(self):
        """创建主窗口控件"""
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ---------------------- 左侧：测试配置和会话选择 ----------------------
        left_container = ttk.Frame(main_frame)
        left_container.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))

        left_canvas = tk.Canvas(left_container, borderwidth=0, highlightthickness=0)
        left_scrollbar = ttk.Scrollbar(left_container, orient=tk.VERTICAL, command=left_canvas.yview)
        left_canvas.pack(side=tk.LEFT, fill=tk.Y, expand=False)
        left_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        left_frame = ttk.Frame(left_canvas)
        left_canvas.create_window((0, 0), window=left_frame, anchor="nw")
        left_canvas.configure(yscrollcommand=left_scrollbar.set)

        def _on_left_mousewheel(event):
            left_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        left_frame.bind("<Enter>", lambda _: left_canvas.bind_all("<MouseWheel>", _on_left_mousewheel))
        left_frame.bind("<Leave>", lambda _: left_canvas.unbind_all("<MouseWheel>"))
        left_frame.bind("<Configure>", lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all")))

        test_manage_frame = ttk.LabelFrame(left_frame, text="测试管理", padding=5)
        test_manage_frame.pack(fill=tk.X, pady=(0, 10))

        # 测试按钮 - 第一组：配置测试和停止测试
        test_btn_frame = ttk.Frame(test_manage_frame)
        test_btn_frame.pack(fill=tk.X, pady=5)

        row1_frame = ttk.Frame(test_btn_frame)
        row1_frame.pack(fill=tk.X, pady=2)

        self.config_btn = ttk.Button(
            row1_frame, text="开始测试", command=self.start_unified_test, width=15
        )
        self.config_btn.pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)
        
        self.stop_btn = ttk.Button(
            row1_frame, text="停止测试", command=self.stop_test, width=15, state=tk.DISABLED
        )
        self.stop_btn.pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        # 第二组：配置管理
        row2_frame = ttk.Frame(test_btn_frame)
        row2_frame.pack(fill=tk.X, pady=2)

        ttk.Button(
            row2_frame, text="配置管理", command=self.open_config_dialog, width=15
        ).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        # 第三组：对比黑名单
        row3_frame = ttk.Frame(test_btn_frame)
        row3_frame.pack(fill=tk.X, pady=2)

        ttk.Button(
            row3_frame, text="对比黑名单", command=self.open_blacklist_compare_dialog, width=15
        ).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        # 第四组：定时任务管理
        row4_frame = ttk.Frame(test_btn_frame)
        row4_frame.pack(fill=tk.X, pady=2)

        ttk.Button(
            row4_frame, text="定时任务", command=self.open_schedule_manager, width=15
        ).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        # 分隔线
        ttk.Separator(left_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)

        # 会话选择 & 结果操作区域已移除（当前版本仅负责执行测试并保存HAR）。

        # 黑名单状态
        blacklist_frame = ttk.LabelFrame(left_frame, text="黑名单拦截", padding=5)
        blacklist_frame.pack(fill=tk.X, pady=5)

        ttk.Label(blacklist_frame, text="当前拦截域名数:").pack(anchor=tk.W)
        self.blacklist_count_var = tk.StringVar(value="0")
        ttk.Label(blacklist_frame, textvariable=self.blacklist_count_var, font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(2, 5))

        # 搜索/添加域名区域
        search_add_frame = ttk.Frame(blacklist_frame)
        search_add_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(search_add_frame, text="搜索/添加:", font=("Arial", 9)).pack(side=tk.LEFT)
        self.blacklist_entry = ttk.Entry(search_add_frame, width=20)
        self.blacklist_entry.pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)
        self.blacklist_entry.bind('<Return>', lambda e: self._on_blacklist_entry_return())
        self.blacklist_entry.bind('<KeyRelease>', lambda e: self._on_blacklist_entry_change())
        ttk.Button(
            search_add_frame, text="添加", command=self.add_domain_to_blacklist, width=8
        ).pack(side=tk.LEFT)

        # 黑名单列表显示
        blacklist_list_frame = ttk.Frame(blacklist_frame)
        blacklist_list_frame.pack(fill=tk.BOTH, expand=True)
        
        list_header_frame = ttk.Frame(blacklist_list_frame)
        list_header_frame.pack(fill=tk.X, pady=(0, 2))
        ttk.Label(list_header_frame, text="拦截列表:", font=("Arial", 9)).pack(side=tk.LEFT)
        ttk.Label(list_header_frame, text="(Ctrl+A全选, Ctrl+C复制)", font=("Arial", 7), foreground="gray").pack(side=tk.RIGHT)
        
        self.blacklist_listbox = tk.Listbox(blacklist_list_frame, height=8, font=("Arial", 8), selectmode=tk.EXTENDED)
        blacklist_scrollbar = ttk.Scrollbar(blacklist_list_frame, orient=tk.VERTICAL, command=self.blacklist_listbox.yview)
        self.blacklist_listbox.configure(yscrollcommand=blacklist_scrollbar.set)
        self.blacklist_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        blacklist_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 绑定快捷键
        self.blacklist_listbox.bind('<Control-a>', lambda e: self._select_all_domains())
        self.blacklist_listbox.bind('<Control-c>', lambda e: self._copy_selected_domains())
        self.blacklist_listbox.bind('<Double-Button-1>', lambda e: self._copy_selected_domains())

        # 存储完整的域名列表（用于搜索过滤）
        self.all_blocked_domains: List[str] = self.blacklist_manager.get_blocked_domains()

        # 黑名单操作按钮
        blacklist_btn_frame = ttk.Frame(blacklist_frame)
        blacklist_btn_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(
            blacklist_btn_frame, text="刷新", command=self.refresh_blacklist_display, width=8
        ).pack(side=tk.LEFT, padx=2)
        ttk.Button(
            blacklist_btn_frame, text="复制选中", command=self._copy_selected_domains, width=10
        ).pack(side=tk.LEFT, padx=2)
        ttk.Button(
            blacklist_btn_frame, text="删除选中", command=self.remove_selected_domain, width=10
        ).pack(side=tk.LEFT, padx=2)
        ttk.Button(
            blacklist_btn_frame, text="清空", command=self.clear_blacklist, width=8
        ).pack(side=tk.LEFT, padx=2)

        # 初始化黑名单显示
        self.refresh_blacklist_display()

        # ---------------------- 右侧：输出目录与测试日志 ----------------------
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # 当前测试输出目录（仅用于查看/打开HAR所在目录）
        output_frame = ttk.LabelFrame(right_frame, text="当前测试输出目录", padding=5)
        output_frame.pack(fill=tk.X, expand=False, pady=(0, 5))

        self.output_dir_var = tk.StringVar(value="暂无测试输出")
        ttk.Label(output_frame, textvariable=self.output_dir_var).pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )
        ttk.Button(
            output_frame,
            text="打开文件夹",
            command=self.open_output_directory,
            width=12,
        ).pack(side=tk.RIGHT, padx=(5, 0))

        # 下部：日志显示
        log_frame = ttk.LabelFrame(right_frame, text="测试日志", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=False, pady=(5, 0))

        # 日志工具栏
        log_toolbar = ttk.Frame(log_frame)
        log_toolbar.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(log_toolbar, text="实时日志:").pack(side=tk.LEFT)

        ttk.Button(
            log_toolbar, text="清空日志", command=self.clear_logs, width=10
        ).pack(side=tk.RIGHT)

        # 日志文本框
        self.log_text = scrolledtext.ScrolledText(
            log_frame, wrap=tk.WORD, height=12
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)

    def clear_logs(self):
        """清空日志显示"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

    def open_output_directory(self):
        """打开当前测试的输出目录（HAR所在目录）"""
        path = self.output_dir_var.get().strip()
        if not path or path == "暂无测试输出":
            messagebox.showinfo("提示", "当前还没有测试输出目录")
            return
        try:
            p = Path(path)
            if not p.exists():
                messagebox.showwarning("提示", f"目录不存在:\n{path}")
                return
            # 在Windows上使用默认文件管理器打开目录
            os.startfile(str(p))
        except Exception as e:
            self.logger.error(f"打开输出目录失败: {e}", exc_info=True)
            messagebox.showerror("错误", f"无法打开目录:\n{path}\n\n错误信息: {e}")

    # ==================== 配置管理相关方法 ====================
    
    def open_config_dialog(self):
        """打开配置管理对话框"""
        def on_config_changed():
            """配置更改后的回调"""
            # 重新设置 Playwright 环境（如果浏览器路径改变了）
            from utils.playwright_path_helper import setup_playwright_environment
            setup_playwright_environment(self.config_manager)
            self.logger.info("配置已更新，已重新设置 Playwright 环境")
        
        dialog = ConfigDialog(self.root, self.config_manager, on_config_changed)
        dialog.show()
    
    def open_schedule_manager(self):
        """打开定时任务管理对话框"""
        show_schedule_manager(self.root, self.scheduler, self.logger, self._execute_scheduled_test)
    
    def _get_output_directory(self, config: Optional[Dict] = None) -> Path:
        """
        获取结果输出目录，优先使用配置中的自定义路径，其次使用全局配置，最后使用默认路径。
        
        Args:
            config: 可选的测试配置字典，如果包含自定义输出目录配置，将优先使用
        
        Returns:
            输出目录的Path对象
        """
        # 1. 优先使用本次测试配置中传递的自定义路径
        if config and config.get('config', {}).get('use_custom_output_dir') and config.get('config', {}).get('custom_output_dir'):
            return Path(config['config']['custom_output_dir'])
        
        # 2. 其次使用全局配置管理器中的自定义路径
        global_custom_path = self.config_manager.get_results_path()
        if global_custom_path:
            return Path(global_custom_path)
        
        # 3. 最后使用默认的results目录
        return get_app_base_dir() / "results"

    # ==================== VPN测试相关方法 ====================


    def _handle_vpn_test_completion(self, session: TestSession, test_name: str, show_message: bool = True):
        """处理单个VPN测试完成"""
        try:
            # 保存结果（仅用于确保HAR目录存在）
            output_dir = self._get_output_directory()
            dir_path = self.tester.save_results(session, output_dir)
            # 记录当前测试输出目录（HAR所在目录）
            self.output_dir_var.set(str(dir_path))
            
        except Exception as e:
            self.logger.error(f"处理VPN测试完成时出错: {e}", exc_info=True)
            if show_message:
                messagebox.showerror("错误", f"处理测试结果时出错:\n{e}")
    
    def _handle_all_vpn_tests_completion(self, completed_sessions: List[tuple], reference_file: str = "", is_scheduled_task: bool = False, test_name: str = ""):
        """处理所有VPN测试完成"""
        try:
            # 结束进度显示（只调用一次）
            self.end_test_progress()
            
            # 如果有按次数执行的测试在等待，通知它们测试完成
            # 根据测试名称匹配事件（测试名称包含"第X次"）
            if test_name and "第" in test_name and "次" in test_name:
                # 尝试找到匹配的事件
                for test_id, event in list(self.test_completion_events.items()):
                    if test_name in test_id and not event.is_set():
                        event.set()
                        self.logger.debug(f"已通知测试完成事件: {test_id} (测试名称: {test_name})")
                        break
            else:
                # 如果没有匹配的测试名称，设置第一个未设置的事件
                for test_id, event in list(self.test_completion_events.items()):
                    if not event.is_set():
                        event.set()
                        self.logger.debug(f"已通知测试完成事件: {test_id}")
                        break
            
            # 处理每个VPN的测试结果（不显示消息框，避免阻塞）
            for session, test_name in completed_sessions:
                self._handle_vpn_test_completion(session, test_name, show_message=False)
            
            # 如果提供了参考文件，进行黑名单对比（定时任务模式下不弹出弹窗）
            if reference_file and not is_scheduled_task:
                # 使用现有的方法进行对比
                for session, _ in completed_sessions:
                    self.compare_and_prompt_blacklist(session, reference_file)
                    break  # 只需要对比一次，因为所有会话的hostname都会合并
            elif reference_file and is_scheduled_task:
                # 定时任务模式下，只记录日志，不弹出弹窗
                self.logger.info(f"定时任务完成，参考文件: {reference_file}，黑名单对比已跳过（定时任务模式下不弹出弹窗）")
            
            # 当前精简模式下，不再弹出"所有VPN测试完成"总结弹窗，仅通过日志记录结果。
            total_vpns = len(completed_sessions)
            self.logger.info(f"所有VPN测试已完成，共 {total_vpns} 个VPN 会话（仅保存HAR文件）")
            
        except Exception as e:
            self.logger.error(f"处理所有VPN测试完成时出错: {e}", exc_info=True)
            # 定时任务模式下不弹出错误弹窗
            if not is_scheduled_task:
                messagebox.showerror("错误", f"处理测试结果时出错:\n{e}")
    
    def _handle_normal_test_completion(self, session: TestSession, test_name: str):
        """处理普通测试完成"""
        try:
            # 结束进度显示
            self.end_test_progress()

            # 保存结果（仅用于确保HAR目录存在）
            output_dir = self._get_output_directory()
            dir_path = self.tester.save_results(session, output_dir)
            # 记录当前测试输出目录（HAR所在目录）
            self.output_dir_var.set(str(dir_path))
            
        except Exception as e:
            self.logger.error(f"处理普通测试完成时出错: {e}")
            messagebox.showerror("错误", f"处理测试结果时出错:\n{e}")
    

    # ==================== 测试进度相关方法 ====================

    def update_test_progress(self, round_num: int, completed_urls: int, total_urls: int, total_rounds: int):
        """更新测试进度显示"""
        self.current_round = round_num
        self.current_url_index = completed_urls
        self.total_urls = total_urls
        self.total_rounds = total_rounds

        # 计算进度百分比
        if total_urls > 0:
            progress_percent = int((completed_urls / total_urls) * 100)
            progress_text = f"测试进度: 第 {round_num + 1}/{total_rounds} 轮 | URL: {completed_urls}/{total_urls} ({progress_percent}%)"
        else:
            progress_text = f"测试进度: 第 {round_num + 1}/{total_rounds} 轮 | URL: {completed_urls}/{total_urls}"
        self.progress_var.set(progress_text)

    def start_test_progress(self, test_name: str, total_urls: int, total_rounds: int):
        """开始测试进度跟踪
        
        注意：is_testing状态应该在调用此方法之前设置，此方法不再设置is_testing状态
        """
        self.test_stopped = False  # 重置停止标志
        self.total_urls = total_urls
        self.total_rounds = total_rounds
        self.current_round = 0
        self.current_url_index = 0

        self.status_var.set(f"正在运行测试 '{test_name}'...")
        self.update_test_progress(0, 0, total_urls, total_rounds)
        self.config_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)  # 启用停止按钮

    def end_test_progress(self):
        """结束测试进度跟踪"""
        with self.testing_lock:
            self.is_testing = False
        self.progress_var.set("")
        self.config_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)  # 禁用停止按钮

    # ==================== 修改测试运行方法 ====================

    def start_unified_test(self):
        """打开统一测试配置对话框并运行测试"""
        # 使用锁保护，防止并发启动多个测试
        with self.testing_lock:
            if self.is_testing:
                messagebox.showinfo("提示", "当前有测试正在运行，请等待测试完成")
                return
            # 立即设置测试状态，防止在对话框显示期间启动其他测试
            self.is_testing = True

        try:
            # 显示统一测试配置对话框
            result = show_unified_test_dialog(self.root, self.config_manager, self.logger)
            
            if result:
                # 更新配置管理器
                self.config_manager.update_config(result['config'])
                
                # 检查是否有定时任务配置
                schedule = result.get('schedule', {})
                if schedule.get('enabled', False):
                    # 添加定时任务
                    from datetime import datetime
                    # 使用ISO 8601基本格式：YYYYMMDDTHHMMSS
                    task_id = f"{result['test_name']}_{datetime.now().strftime('%Y%m%dT%H%M%S')}"
                    
                    # 创建任务配置（包含schedule信息）
                    task_config = result.copy()
                    
                    # 添加定时任务到调度器
                    self.scheduler.add_task(
                        task_id=task_id,
                        config=task_config,
                        callback=self._execute_scheduled_test
                    )
                    
                    # 显示提示
                    weekday_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
                    selected_days = [weekday_names[day] for day in schedule.get('weekdays', [])]
                    times_str = ', '.join(schedule.get('times', []))
                    
                    messagebox.showinfo(
                        "定时任务已添加",
                        f"定时任务已添加！\n\n"
                        f"任务名称: {result['test_name']}\n"
                        f"执行日期: {', '.join(selected_days)}\n"
                        f"执行时间: {times_str}\n\n"
                        f"程序将在指定时间自动执行测试。"
                    )
                    # 定时任务不需要立即设置is_testing状态（会在执行时设置）
                    with self.testing_lock:
                        self.is_testing = False
                else:
                    # 检查是否是按次数执行模式
                    test_mode = result.get('test_mode', 'single')
                    if test_mode == 'repeat':
                        # 按次数执行模式
                        repeat_count = result.get('repeat_count', 1)
                        start_time_mode = result.get('start_time_mode', 'immediate')
                        start_datetime = result.get('start_datetime')
                        
                        if start_time_mode == 'scheduled' and start_datetime:
                            # 定时开始：等待到指定时间
                            self._run_repeat_test_scheduled(result, repeat_count, start_datetime)
                        else:
                            # 立即开始：直接执行
                            self._run_repeat_test_immediate(result, repeat_count)
                    else:
                        # 单次执行模式（is_testing已在start_unified_test中设置）
                        self.run_vpn_test_with_config(result)
            else:
                # 用户取消了对话框，清除测试状态
                with self.testing_lock:
                    self.is_testing = False
        except Exception as e:
            # 如果出现异常，清除测试状态
            with self.testing_lock:
                self.is_testing = False
            self.logger.error(f"启动测试失败: {e}", exc_info=True)
            messagebox.showerror("错误", f"启动测试失败: {e}")
    
    def _run_repeat_test_immediate(self, config: Dict, repeat_count: int):
        """立即开始按次数执行测试"""
        import threading
        from datetime import datetime
        
        def run_repeat_loop():
            try:
                for i in range(repeat_count):
                    self.logger.info(f"开始第 {i+1}/{repeat_count} 次测试")
                    # 更新测试名称，添加次数标识
                    test_config = config.copy()
                    test_config['test_name'] = f"{config['test_name']}_第{i+1}次"
                    
                    # 创建测试完成事件
                    test_id = f"{test_config['test_name']}_{i}"
                    test_completed = threading.Event()
                    self.test_completion_events[test_id] = test_completed
                    
                    try:
                        # 在主线程中执行测试（is_testing状态已在start_unified_test中设置）
                        self.root.after(0, lambda c=test_config: self.run_vpn_test_with_config(c))
                        
                        # 等待测试完成（等待事件）
                        # 设置超时时间（30分钟），避免无限等待
                        if not test_completed.wait(timeout=1800):
                            self.logger.warning(f"第 {i+1} 次测试等待超时（30分钟）")
                            break
                        
                        # 如果不是最后一次，等待一下再执行下一次
                        if i < repeat_count - 1:
                            self.logger.info(f"第 {i+1} 次测试完成，等待5秒后开始第 {i+2} 次测试...")
                            threading.Event().wait(5)
                            # 准备下一次测试，重新设置测试状态
                            with self.testing_lock:
                                self.is_testing = True
                    finally:
                        # 清理事件
                        self.test_completion_events.pop(test_id, None)
                
                self.logger.info(f"所有 {repeat_count} 次测试已完成")
                # 所有测试完成后，清除测试状态
                with self.testing_lock:
                    self.is_testing = False
            except Exception as e:
                self.logger.error(f"按次数执行测试失败: {e}", exc_info=True)
                # 出现异常时清除测试状态
                with self.testing_lock:
                    self.is_testing = False
                # 清理所有事件
                self.test_completion_events.clear()
        
        # 在后台线程中执行循环
        threading.Thread(target=run_repeat_loop, daemon=True).start()
    
    def _run_repeat_test_scheduled(self, config: Dict, repeat_count: int, start_datetime):
        """定时开始按次数执行测试"""
        import threading
        from datetime import datetime
        
        def wait_and_run():
            try:
                now = datetime.now()
                if start_datetime > now:
                    wait_seconds = (start_datetime - now).total_seconds()
                    self.logger.info(f"等待到 {start_datetime.strftime('%Y-%m-%d %H:%M:%S')} 开始执行（还需等待 {wait_seconds:.0f} 秒）")
                    
                    # 显示等待提示
                    self.root.after(0, lambda: self.progress_var.set(
                        f"等待开始时间: {start_datetime.strftime('%Y-%m-%d %H:%M:%S')} (还需等待 {wait_seconds:.0f} 秒)"
                    ))
                    
                    # 等待到开始时间
                    threading.Event().wait(wait_seconds)
                
                # 开始执行循环
                self.logger.info(f"到达开始时间，开始执行 {repeat_count} 次测试")
                self._run_repeat_test_immediate(config, repeat_count)
            except Exception as e:
                self.logger.error(f"定时开始按次数执行测试失败: {e}", exc_info=True)
        
        # 在后台线程中等待并执行
        threading.Thread(target=wait_and_run, daemon=True).start()
    
    def _execute_scheduled_test(self, config: Dict):
        """执行定时任务"""
        try:
            task_name = config.get('test_name', '未知')
            self.logger.info(f"执行定时任务: {task_name}")
            
            # 检查是否有测试正在运行
            with self.testing_lock:
                if self.is_testing:
                    self.logger.warning(f"定时任务 '{task_name}' 无法执行，当前有测试正在运行，将跳过本次执行")
                    return
                # 立即设置测试状态，防止与其他测试并发
                self.is_testing = True
            
            # 为定时任务创建独立的文件夹结构
            # 格式：results/定时任务名称/VPN名称_时间戳/
            from datetime import datetime
            from utils.file_utils import get_app_base_dir
            
            # 获取输出目录（优先使用自定义路径）
            base_output_dir = self._get_output_directory(config)
            
            # 创建定时任务的基础目录（在自定义路径下创建任务名称文件夹）
            task_base_dir = base_output_dir / task_name
            task_base_dir.mkdir(parents=True, exist_ok=True)
            
            # 修改配置，指定输出目录
            config_with_dir = config.copy()
            config_with_dir['_scheduled_task_base_dir'] = str(task_base_dir)
            config_with_dir['_is_scheduled_task'] = True  # 标记为定时任务
            
            # 在主线程中执行测试（is_testing状态已在上面的锁中设置）
            self.root.after(0, lambda: self.run_vpn_test_with_config_scheduled(config_with_dir))
        except Exception as e:
            self.logger.error(f"执行定时任务失败: {e}", exc_info=True)
            # 出现异常时清除测试状态
            with self.testing_lock:
                self.is_testing = False
    
    def run_vpn_test_with_config_scheduled(self, config: Dict):
        """运行定时任务的VPN测试（使用独立的文件夹结构）"""
        try:
            # 测试状态已在_execute_scheduled_test中设置，这里不需要重复设置
            self.start_test_progress(config['test_name'], len(config['urls']), 1)
            
            # 创建测试器
            self.tester = PageLoadTester(
                logger=self.logger,
                config_manager=self.config_manager
            )
            
            # 配置测试器选项
            self.tester.max_concurrent = config['config']['max_concurrent_tests']
            self.tester.stage_timeouts = {
                'stage1': config['config']['stage_timeouts']['stage1'] * 1000,  # 转换为毫秒
                'stage2': config['config']['stage_timeouts']['stage2'] * 1000,
                'stage3': config['config']['stage_timeouts']['stage3'] * 1000
            }
            # 配置HAR选项（包括extract_hostnames）
            self.tester.har_options = {
                "enable_har_capture": config['config']['enable_har_capture'],
                "extract_hostnames": config['config'].get('enable_hostname_capture', True),
                "save_har_files": True,
                "max_har_size_mb": 0  # 不限制大小
            }
            
            # 传递黑名单管理器
            self.tester.blacklist_manager = self.blacklist_manager
            
            # 初始化VPN支持
            self.tester.initialize_vpn_support(True)
            
            # 获取定时任务的基础目录（优先使用配置中的自定义路径）
            if config.get('_scheduled_task_base_dir'):
                task_base_dir = Path(config['_scheduled_task_base_dir'])
            else:
                # 如果没有指定，使用自定义路径或默认路径
                task_base_dir = self._get_output_directory(config) / config.get('test_name', 'ScheduledTask')
            
            # 启动多VPN测试线程
            def run_multiple_vpn_test_thread():
                try:
                    completed_sessions = []  # 存储所有完成的会话
                    
                    # 对VPN列表进行排序，确保每次执行顺序一致
                    sorted_vpns = sorted(config['selected_vpns'])
                    self.logger.info(f"VPN列表已排序: {sorted_vpns}")
                    
                    # 为每个VPN单独执行测试
                    for i, vpn_name in enumerate(sorted_vpns):
                        from datetime import datetime
                        from core.tester import PageLoadTester
                        # 使用ISO 8601基本格式：YYYYMMDDTHHMMSS
                        timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
                        # 转换VPN名称为英文
                        vpn_english = PageLoadTester._convert_vpn_name_to_english(vpn_name)
                        individual_test_name = f"{vpn_english}_{timestamp}"
                        
                        # 为每个VPN创建独立的子目录（转换为英文名称）
                        from core.tester import PageLoadTester
                        vpn_english_name = PageLoadTester._convert_vpn_name_to_english(vpn_name)
                        vpn_dir = task_base_dir / vpn_english_name
                        vpn_dir.mkdir(parents=True, exist_ok=True)
                        
                        self.logger.info(f"开始执行VPN测试 ({i+1}/{len(sorted_vpns)}): {vpn_name}")
                        
                        # 更新进度：开始测试VPN
                        total_vpns = len(sorted_vpns)
                        vpn_progress_text = f"VPN进度: {i+1}/{total_vpns} ({vpn_name}) | URL: 0/{len(config['urls'])}"
                        self.root.after(0, lambda t=vpn_progress_text: self.progress_var.set(t))
                        
                        # 为单个VPN执行测试（指定VPN子目录作为输出目录）
                        # 需要传递进度更新回调
                        session = self.tester.run_vpn_tests(
                            urls=config['urls'],
                            vpn_names=[vpn_name],  # 只测试当前VPN
                            test_name=individual_test_name,
                            rounds=1,
                            concurrency=config['config']['max_concurrent_tests'],
                            base_output_dir=vpn_dir,  # 指定VPN子目录作为输出目录
                            progress_callback=lambda completed, total: self.root.after(0, 
                                lambda c=completed, t=total, vpn=vpn_name, vpn_idx=i+1, vpn_total=total_vpns: 
                                self.progress_var.set(f"VPN进度: {vpn_idx}/{vpn_total} ({vpn}) | URL: {c}/{t} ({int((c/t)*100) if t > 0 else 0}%)"))
                        )
                        
                        # 保存结果到VPN子目录
                        self.tester.save_results(session, vpn_dir)
                        
                        # 保存会话信息（不立即处理UI，避免阻塞）
                        completed_sessions.append((session, individual_test_name))
                        
                        # 如果不是最后一个VPN，等待一段时间
                        if i < len(sorted_vpns) - 1:
                            import time
                            time.sleep(2)  # 等待2秒再测试下一个VPN
                    
                    # 所有VPN测试完成，在主线程中统一处理
                    self.logger.info("所有VPN测试已完成")
                    reference_file = config.get('reference_file', '')
                    current_test_name = config.get('test_name', '')
                    # 定时任务模式下不弹出弹窗
                    is_scheduled = config.get('_is_scheduled_task', False)
                    self.root.after(0, lambda sessions=completed_sessions, ref=reference_file, scheduled=is_scheduled, name=current_test_name: self._handle_all_vpn_tests_completion(sessions, ref, scheduled, name))
                    
                except Exception as e:
                    error_msg = f"VPN测试执行失败: {e}"
                    self.logger.error(error_msg, exc_info=True)
                    self.root.after(0, lambda msg=error_msg: self._handle_test_error(msg))
            
            self._start_managed_thread(run_multiple_vpn_test_thread, name="vpn-scheduled-thread")
            
        except Exception as e:
            self.logger.error(f"启动VPN测试失败: {e}")
            # 清除测试状态
            with self.testing_lock:
                self.is_testing = False
            self._handle_test_error(f"启动测试失败: {e}")

    def run_normal_test_with_config(self, config):
        """运行普通测试"""
        try:
            # 检查是否有测试正在运行
            with self.testing_lock:
                if self.is_testing:
                    self.logger.warning("无法启动普通测试，当前有测试正在运行")
                    messagebox.showwarning("警告", "当前有测试正在运行，请等待测试完成")
                    return
                self.is_testing = True
            
            self.start_test_progress(config['test_name'], len(config['urls']), 1)
            
            # 创建测试器
            self.tester = PageLoadTester(
                logger=self.logger,
                config_manager=self.config_manager
            )
            
            # 配置测试器选项
            self.tester.max_concurrent = config['config']['max_concurrent_tests']
            self.tester.stage_timeouts = {
                'stage1': config['config']['stage_timeouts']['stage1'] * 1000,  # 转换为毫秒
                'stage2': config['config']['stage_timeouts']['stage2'] * 1000,
                'stage3': config['config']['stage_timeouts']['stage3'] * 1000
            }
            # 配置HAR选项（包括extract_hostnames）
            self.tester.har_options = {
                "enable_har_capture": config['config']['enable_har_capture'],
                "extract_hostnames": config['config'].get('enable_hostname_capture', True),
                "save_har_files": True,
                "max_har_size_mb": 0  # 不限制大小
            }
            
            # 传递黑名单管理器
            self.tester.blacklist_manager = self.blacklist_manager
            
            # 启动测试线程
            def run_test_thread():
                try:
                    session = self.tester.run_tests(
                        urls=config['urls'],
                        test_name=config['test_name'],
                        rounds=1  # 固定为1轮
                    )
                    self.root.after(0, lambda s=session, n=config['test_name']: self._handle_normal_test_completion(s, n))
                except Exception as e:
                    error_msg = f"测试执行失败: {e}"
                    # 如果已经请求停止（窗口关闭），不再调度UI线程，避免主线程已退出
                    if self.test_stopped:
                        self.logger.info(f"测试被终止：{error_msg}")
                        return
                    self.logger.error(f"普通测试执行失败: {e}")
                    self._safe_after(0, lambda msg=error_msg: self._handle_test_error(msg))
            
            self._start_managed_thread(run_test_thread, name="normal-test-thread")
            
        except Exception as e:
            self.logger.error(f"启动普通测试失败: {e}")
            self._handle_test_error(f"启动测试失败: {e}")
    
    def run_normal_test_with_config_unified(self, config):
        """运行普通测试（用于统一测试对话框，不使用VPN）"""
        try:
            # 检查是否有测试正在运行
            with self.testing_lock:
                if self.is_testing:
                    self.logger.warning("无法启动普通测试，当前有测试正在运行")
                    messagebox.showwarning("警告", "当前有测试正在运行，请等待测试完成")
                    return
                self.is_testing = True
            
            self.start_test_progress(config['test_name'], len(config['urls']), 1)
            
            # 创建测试器
            self.tester = PageLoadTester(
                logger=self.logger,
                config_manager=self.config_manager
            )
            
            # 配置测试器选项
            self.tester.max_concurrent = config['config']['max_concurrent_tests']
            self.tester.stage_timeouts = {
                'stage1': config['config']['stage_timeouts']['stage1'] * 1000,  # 转换为毫秒
                'stage2': config['config']['stage_timeouts']['stage2'] * 1000,
                'stage3': config['config']['stage_timeouts']['stage3'] * 1000
            }
            self.tester.har_options = {
                "enable_har_capture": config['config']['enable_har_capture'],
                "extract_hostnames": config['config'].get('enable_hostname_capture', True),
                "save_har_files": True,
                "max_har_size_mb": 0  # 不限制大小
            }
            self.tester.headless = config['config'].get('headless', True)
            self.tester.wait_for_network_idle = config['config'].get('wait_for_network_idle', False)
            
            # 传递黑名单管理器
            self.tester.blacklist_manager = self.blacklist_manager
            
            # 不初始化VPN支持（使用当前网络环境）
            self.tester.initialize_vpn_support(False)
            
            # 启动测试线程
            def run_test_thread():
                try:
                    session = self.tester.run_tests(
                        urls=config['urls'],
                        test_name=config['test_name'],
                        rounds=1,  # 固定为1轮
                        concurrency=config['config']['max_concurrent_tests']
                    )
                    
                    # 保存结果并处理完成
                    from utils.file_utils import get_app_base_dir
                    output_dir = self._get_output_directory()
                    filename = self.tester.save_results(session, output_dir)
                    reference_file = config.get('reference_file', '')
                    self.root.after(0, lambda s=session, n=config['test_name'], f=filename, ref=reference_file: 
                                  self._handle_test_completion(s, n, f, ref))
                except Exception as e:
                    error_msg = f"测试执行失败: {e}"
                    if self.test_stopped:
                        self.logger.info(f"测试被终止：{error_msg}")
                        return
                    self.logger.error(f"普通测试执行失败: {e}", exc_info=True)
                    self._safe_after(0, lambda msg=error_msg: self._handle_test_error(msg))
            
            self._start_managed_thread(run_test_thread, name="unified-test-thread")
            
        except Exception as e:
            self.logger.error(f"启动普通测试失败: {e}", exc_info=True)
            # 清除测试状态
            with self.testing_lock:
                self.is_testing = False
            self._handle_test_error(f"启动测试失败: {e}")
    
    def run_direct_only_test_with_config(self, config):
        """运行直连模式测试（不连接VPN，只测试直连通道）"""
        try:
            selected_vpns = config.get('selected_vpns', [])
            if not selected_vpns:
                self.logger.warning("直连模式需要选择至少一个VPN作为标识，将使用普通测试")
                self.run_normal_test_with_config_unified(config)
                return
            
            # 检查是否有测试正在运行
            with self.testing_lock:
                if self.is_testing:
                    self.logger.warning("无法启动直连模式测试，当前有测试正在运行")
                    messagebox.showwarning("警告", "当前有测试正在运行，请等待测试完成")
                    return
                self.is_testing = True
            
            self.start_test_progress(config['test_name'], len(config['urls']), 1)
            
            # 创建测试器
            self.tester = PageLoadTester(
                logger=self.logger,
                config_manager=self.config_manager
            )
            
            # 配置测试器选项
            self.tester.max_concurrent = config['config']['max_concurrent_tests']
            self.tester.stage_timeouts = {
                'stage1': config['config']['stage_timeouts']['stage1'] * 1000,  # 转换为毫秒
                'stage2': config['config']['stage_timeouts']['stage2'] * 1000,
                'stage3': config['config']['stage_timeouts']['stage3'] * 1000
            }
            self.tester.har_options = {
                "enable_har_capture": config['config']['enable_har_capture'],
                "extract_hostnames": config['config'].get('enable_hostname_capture', True),
                "save_har_files": True,
                "max_har_size_mb": 0  # 不限制大小
            }
            self.tester.headless = config['config'].get('headless', True)
            self.tester.wait_for_network_idle = config['config'].get('wait_for_network_idle', False)
            
            # 传递黑名单管理器
            self.tester.blacklist_manager = self.blacklist_manager
            
            # 不初始化VPN支持（使用当前网络环境，不连接VPN）
            self.tester.initialize_vpn_support(False)
            
            # 获取输出目录（优先使用自定义路径）
            base_output_dir = self._get_output_directory(config)
            
            # 启动测试线程
            def run_direct_test_thread():
                try:
                    completed_sessions = []  # 存储所有完成的会话
                    
                    # 对VPN列表进行排序，确保每次执行顺序一致
                    sorted_vpns = sorted(selected_vpns)
                    self.logger.info(f"直连模式测试，VPN列表已排序: {sorted_vpns}")
                    
                    # 为每个VPN单独执行测试（但不连接VPN，只测试直连）
                    for i, vpn_name in enumerate(sorted_vpns):
                        from datetime import datetime
                        from core.tester import PageLoadTester
                        # 使用ISO 8601基本格式：YYYYMMDDTHHMMSS
                        timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
                        # 转换VPN名称为英文
                        vpn_english = PageLoadTester._convert_vpn_name_to_english(vpn_name)
                        individual_test_name = f"{vpn_english}_direct_{timestamp}"
                        
                        # 为每个VPN创建独立的子目录
                        vpn_english_name = PageLoadTester._convert_vpn_name_to_english(vpn_name)
                        vpn_dir = base_output_dir / vpn_english_name
                        vpn_dir.mkdir(parents=True, exist_ok=True)
                        
                        self.logger.info(f"开始执行直连模式测试 ({i+1}/{len(sorted_vpns)}): {vpn_name}（不连接VPN）")
                        
                        # 更新进度：开始测试直连
                        total_vpns = len(sorted_vpns)
                        vpn_progress_text = f"直连进度: {i+1}/{total_vpns} ({vpn_name}) | URL: 0/{len(config['urls'])}"
                        self.root.after(0, lambda t=vpn_progress_text: self.progress_var.set(t))
                        
                        # 执行直连模式测试（不连接VPN）
                        session = self.tester.run_direct_only_tests(
                            urls=config['urls'],
                            vpn_name=vpn_name,  # 用于标识和文件命名
                            test_name=individual_test_name,
                            rounds=1,
                            concurrency=config['config']['max_concurrent_tests'],
                            base_output_dir=vpn_dir,  # 使用VPN子目录作为输出目录
                            progress_callback=lambda completed, total: self.root.after(0, 
                                lambda c=completed, t=total, vpn=vpn_name, vpn_idx=i+1, vpn_total=total_vpns: 
                                self.progress_var.set(f"直连进度: {vpn_idx}/{vpn_total} ({vpn}) | URL: {c}/{t} ({int((c/t)*100) if t > 0 else 0}%)"))
                        )
                        
                        # 保存结果到VPN子目录
                        self.tester.save_results(session, vpn_dir)
                        
                        # 保存会话信息（不立即处理UI，避免阻塞）
                        completed_sessions.append((session, individual_test_name))
                        
                        # 如果不是最后一个VPN，等待一段时间
                        if i < len(sorted_vpns) - 1:
                            import time
                            time.sleep(1)  # 等待1秒再测试下一个VPN
                    
                    # 所有测试完成，在主线程中统一处理
                    self.logger.info("所有直连模式测试已完成")
                    reference_file = config.get('reference_file', '')
                    current_test_name = config.get('test_name', '')
                    self.root.after(0, lambda sessions=completed_sessions, ref=reference_file, name=current_test_name: self._handle_all_vpn_tests_completion(sessions, ref, False, name))
                    
                except Exception as e:
                    error_msg = f"直连模式测试执行失败: {e}"
                    self.logger.error(error_msg, exc_info=True)
                    self.root.after(0, lambda msg=error_msg: self._handle_test_error(msg))
            
            self._start_managed_thread(run_direct_test_thread, name="direct-test-thread")
            
        except Exception as e:
            self.logger.error(f"启动直连模式测试失败: {e}")
            # 清除测试状态
            with self.testing_lock:
                self.is_testing = False
            self._handle_test_error(f"启动测试失败: {e}")
    
    def run_vpn_test_with_config(self, config):
        """运行测试（支持VPN测试和普通测试）"""
        try:
            selected_vpns = config.get('selected_vpns', [])
            enable_vpn_test = config.get('enable_vpn_test', True)
            enable_direct_test = config.get('enable_direct_test', True)
            
            # 如果没有选择VPN，使用普通测试（直接使用当前网络环境）
            if not selected_vpns:
                self.logger.info("未选择VPN，使用当前网络环境进行普通测试")
                self.run_normal_test_with_config_unified(config)
                return
            
            # 如果只启用了直连测试，不连接VPN，只测试直连
            if not enable_vpn_test and enable_direct_test:
                self.logger.info(f"只启用直连模式测试，将只测试直连通道（不连接VPN），VPN列表: {selected_vpns}")
                self.run_direct_only_test_with_config(config)
                return
            
            # 如果选择了VPN，使用VPN测试（根据配置决定测试VPN模式、直连模式或两者）
            self.logger.info(f"已选择VPN: {selected_vpns}，VPN模式测试: {enable_vpn_test}，直连模式测试: {enable_direct_test}")
            
            # 测试状态已在start_unified_test中设置（单次执行模式）
            # 但如果是从其他地方调用的（如按次数执行模式），需要设置
            with self.testing_lock:
                if not self.is_testing:
                    self.is_testing = True
            
            self.start_test_progress(config['test_name'], len(config['urls']), 1)
            
            # 创建测试器
            self.tester = PageLoadTester(
                logger=self.logger,
                config_manager=self.config_manager
            )
            
            # 配置测试器选项
            self.tester.max_concurrent = config['config']['max_concurrent_tests']
            self.tester.stage_timeouts = {
                'stage1': config['config']['stage_timeouts']['stage1'] * 1000,  # 转换为毫秒
                'stage2': config['config']['stage_timeouts']['stage2'] * 1000,
                'stage3': config['config']['stage_timeouts']['stage3'] * 1000
            }
            self.tester.har_options = {
                "enable_har_capture": config['config']['enable_har_capture'],
                "enable_hostname_capture": config['config']['enable_hostname_capture']
            }
            
            # 传递黑名单管理器
            self.tester.blacklist_manager = self.blacklist_manager
            
            # 初始化VPN支持
            self.tester.initialize_vpn_support(True)
            
            # 启动多VPN测试线程
            def run_multiple_vpn_test_thread():
                try:
                    completed_sessions = []  # 存储所有完成的会话
                    
                    # 对VPN列表进行排序，确保每次执行顺序一致
                    sorted_vpns = sorted(config['selected_vpns'])
                    self.logger.info(f"VPN列表已排序: {sorted_vpns}")
                    
                    # 为每个VPN单独执行测试
                    for i, vpn_name in enumerate(sorted_vpns):
                        from datetime import datetime
                        from core.tester import PageLoadTester
                        # 使用ISO 8601基本格式：YYYYMMDDTHHMMSS
                        timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
                        # 转换VPN名称为英文
                        vpn_english = PageLoadTester._convert_vpn_name_to_english(vpn_name)
                        individual_test_name = f"{vpn_english}_{timestamp}"
                        
                        self.logger.info(f"开始执行VPN测试 ({i+1}/{len(sorted_vpns)}): {vpn_name}")
                        
                        # 更新进度：开始测试VPN
                        total_vpns = len(sorted_vpns)
                        vpn_progress_text = f"VPN进度: {i+1}/{total_vpns} ({vpn_name}) | URL: 0/{len(config['urls'])}"
                        self.root.after(0, lambda t=vpn_progress_text: self.progress_var.set(t))
                        
                        # 为单个VPN执行测试
                        session = self.tester.run_vpn_tests(
                            urls=config['urls'],
                            vpn_names=[vpn_name],  # 只测试当前VPN
                            test_name=individual_test_name,
                            rounds=1,
                            concurrency=config['config']['max_concurrent_tests'],
                            progress_callback=lambda completed, total: self.root.after(0, 
                                lambda c=completed, t=total, vpn=vpn_name, vpn_idx=i+1, vpn_total=total_vpns: 
                                self.progress_var.set(f"VPN进度: {vpn_idx}/{vpn_total} ({vpn}) | URL: {c}/{t} ({int((c/t)*100) if t > 0 else 0}%)")),
                            enable_vpn_test=config.get('enable_vpn_test', True),
                            enable_direct_test=config.get('enable_direct_test', True)
                        )
                        
                        # 保存会话信息（不立即处理UI，避免阻塞）
                        completed_sessions.append((session, individual_test_name))
                        
                        # 如果不是最后一个VPN，等待一段时间
                        if i < len(sorted_vpns) - 1:
                            import time
                            time.sleep(2)  # 等待2秒再测试下一个VPN
                    
                    # 所有VPN测试完成，在主线程中统一处理
                    self.logger.info("所有VPN测试已完成")
                    reference_file = config.get('reference_file', '')
                    current_test_name = config.get('test_name', '')
                    self.root.after(0, lambda sessions=completed_sessions, ref=reference_file, name=current_test_name: self._handle_all_vpn_tests_completion(sessions, ref, False, name))
                    
                except Exception as e:
                    error_msg = f"VPN测试执行失败: {e}"
                    self.logger.error(error_msg, exc_info=True)
                    self.root.after(0, lambda msg=error_msg: self._handle_test_error(msg))
            
            self._start_managed_thread(run_multiple_vpn_test_thread, name="vpn-test-thread")
            
        except Exception as e:
            self.logger.error(f"启动VPN测试失败: {e}")
            self._handle_test_error(f"启动测试失败: {e}")
    
    def run_test_with_config(self, config):
        """使用配置运行测试"""
        # 检查对比文件是否存在
        reference_file = config.get('reference_file', '')
        if reference_file:
            ref_path = Path(reference_file)
            if not ref_path.exists():
                self.logger.warning(f"对比文件不存在: {reference_file}")
                messagebox.showwarning("文件缺失", f"对比文件不存在，将跳过对比分析:\n{reference_file}\n\n请检查文件路径是否正确。")
        
        # 开始进度跟踪
        self.start_test_progress(config['test_name'], len(config['urls']), config['rounds'])

        def run_in_thread():
            try:
                stage_timeouts_sec = config.get('stage_timeouts') or {}
                stage_timeouts_ms = {
                    'stage1': int(stage_timeouts_sec.get('stage1', 8)) * 1000,
                    'stage2': int(stage_timeouts_sec.get('stage2', 15)) * 1000,
                    'stage3': int(stage_timeouts_sec.get('stage3', 30)) * 1000,
                }
                self.tester.stage_timeouts = stage_timeouts_ms
                # 兼容旧逻辑：仍然设置整体超时为阶段3
                overall_timeout = config.get('timeout', stage_timeouts_sec.get('stage3', 30))
                self.tester.timeout = int(overall_timeout) * 1000
                self.tester.headless = config['headless']
                self.tester.wait_for_network_idle = config['wait_for_network_idle']
                self.tester.update_har_options(config.get('har_options'))

                # 修改测试器以支持进度回调
                test_session = self.run_tests_with_progress(
                    config['urls'],
                    config['concurrency'],
                    config['rounds'],
                    config['test_name']
                )

                # 获取输出目录（优先使用自定义路径）
                output_dir = self._get_output_directory(config)
                filename = self.tester.save_results(test_session, output_dir)
                reference_file = config.get('reference_file', '')
                self.root.after(0, self._handle_test_completion, test_session, test_session.test_name, filename, reference_file)

            except Exception as e:
                self.root.after(0, self._handle_test_error, str(e))

        # 启动线程
        self._start_managed_thread(run_in_thread, name="config-test-thread")

    def run_tests_with_progress(self, urls: List[str], concurrency: int, rounds: int, test_name: str) -> TestSession:
        """
        运行批量测试，支持进度回调

        注意：这个方法需要在后台线程中运行
        """
        # 创建带时间戳的测试名称（用于区分不同会话）
        # 使用ISO 8601基本格式：YYYYMMDDTHHMMSS
        timestamp_suffix = datetime.now().strftime("%Y%m%dT%H%M%S")
        base_name = test_name.strip() if test_name else "Test"
        session_test_name = f"{base_name}_{timestamp_suffix}"

        # 创建测试会话
        session = TestSession(
            test_name=session_test_name,
            test_rounds=rounds,
            total_urls=len(urls)
        )
        # 获取输出目录（优先使用自定义路径）
        output_dir = self._get_output_directory()
        self.tester.prepare_session_directory(session, output_dir)

        self.logger.info(f"开始测试 '{session.test_name}': {len(urls)} 个URL (并发数: {concurrency}, 轮次: {rounds})")

        # 保存当前会话引用（用于停止时保存）
        self.current_test_session = session
        
        # 使用线程池实现真正的并发
        actual_concurrency = min(concurrency, len(urls))  # 实际并发数不超过URL数量

        # 多轮次测试
        for round_num in range(rounds):
            # 检查是否已停止
            if self.test_stopped:
                self.logger.info("测试已被用户停止")
                break
            
            # 每轮测试开始前统一清理一次
            self.tester._clear_all_caches(f"第 {round_num + 1} 轮测试开始前")
                
            self.logger.info(f"=== 开始第 {round_num + 1}/{rounds} 轮测试 ===")

            # 更新进度：新的一轮开始
            self.root.after(0, self.update_test_progress, round_num, 0, len(urls), rounds)

            # 准备所有测试任务队列
            from queue import Queue
            task_queue = Queue()
            for url_index, url in enumerate(urls):
                original_position = url_index + 1
                task_queue.put((url, url_index, original_position, round_num))

            # 使用线程池动态调度执行
            completed_count = 0
            total_tasks = len(urls)
            future_to_task = {}
            
            with ThreadPoolExecutor(max_workers=actual_concurrency) as executor:
                # 初始提交：提交与并发数相等的任务
                for _ in range(min(actual_concurrency, total_tasks)):
                    if not task_queue.empty():
                        url, url_index, original_position, _ = task_queue.get()
                        future = executor.submit(
                            self._run_single_test, url, url_index, original_position, round_num
                        )
                        future_to_task[future] = (url, url_index, original_position)

                # 动态调度：当任务完成时，立即提交下一个任务
                while future_to_task and not self.test_stopped:
                    # 等待任意一个任务完成
                    done, not_done = wait(future_to_task.keys(), return_when=FIRST_COMPLETED)
                    
                    # 处理完成的任务
                    for future in done:
                        url, url_index, original_position = future_to_task.pop(future)
                        completed_count += 1

                        try:
                            result = future.result()
                            # 处理结果，按URL分组存储
                            if result.url not in session.results:
                                session.results[result.url] = []
                            session.results[result.url].append(result)

                            # 更新进度
                            self.root.after(0, self.update_test_progress, round_num, completed_count - 1, len(urls), rounds)

                        except Exception as exc:
                            self.logger.error(f"测试URL [{original_position}-{round_num}] {url} 时发生异常: {exc}", exc_info=True)
                            # 创建错误结果
                            from core.result import TestResult
                            error_result = TestResult(
                                url=url,
                                status="error",
                                error_type="exception",
                                error_message=str(exc),
                                test_round=round_num,
                                url_index=url_index,
                                original_position=original_position,
                                response_time=-1
                            )
                            if error_result.url not in session.results:
                                session.results[error_result.url] = []
                            session.results[error_result.url].append(error_result)

                        # 立即提交下一个任务（如果有且未停止）
                        if not task_queue.empty() and not self.test_stopped:
                            next_url, next_url_index, next_original_position, _ = task_queue.get()
                            next_future = executor.submit(
                                self._run_single_test, next_url, next_url_index, next_original_position, round_num
                            )
                            future_to_task[next_future] = (next_url, next_url_index, next_original_position)
                    
                    # 如果已停止，取消所有未完成的任务
                    if self.test_stopped:
                        for future in list(future_to_task.keys()):
                            future.cancel()
                        future_to_task.clear()
                        break

            # 检查是否已停止
            if self.test_stopped:
                self.logger.info(f"测试在第 {round_num + 1} 轮被停止")
                break
            
            self.logger.info(f"第 {round_num + 1}/{rounds} 轮测试完成")

            # 轮次之间暂停（可选）
            if round_num < rounds - 1 and not self.test_stopped:
                self.logger.info(f"等待 {5} 秒后开始下一轮测试...")
                # 在暂停期间也检查停止标志
                for _ in range(50):  # 5秒 = 50 * 0.1秒
                    if self.test_stopped:
                        break
                    time.sleep(0.1)

        # 如果测试被停止，询问是否保存
        if self.test_stopped:
            self.logger.info("测试已停止，准备处理已测试的结果...")
            # 在主线程中显示保存对话框
            self.root.after(0, self._handle_stopped_test, session)
            return session
        
        # 更新HAR相关统计再计算元数据
        self.tester.apply_har_metadata(session)
        session.calculate_metadata()

        self.logger.info(
            f"测试 '{session.test_name}' 完成: "
            f"总请求 {session.metadata['total_requests']}, "
            f"成功 {session.metadata['successful_requests']}, "
            f"成功率 {session.metadata['success_rate']}%"
        )

        return session
    
    def stop_test(self):
        """停止当前测试"""
        if not self.is_testing:
            return
        
        # 确认停止
        if messagebox.askyesno("停止测试", "确定要停止当前测试吗？"):
            self.test_stopped = True
            self.logger.info("用户请求停止测试...")
            self.status_var.set("正在停止测试...")
    
    def _handle_stopped_test(self, session: TestSession):
        """处理停止的测试"""
        # 计算已测试的结果数量
        total_tested = sum(len(results) for results in session.results.values())
        
        if total_tested == 0:
            messagebox.showinfo("提示", "没有已测试的结果需要保存")
            self.end_test_progress()
            return
        
        # 询问是否保存
        msg = f"测试已停止，共有 {total_tested} 个测试结果。\n\n是否保存已测试的结果？"
        if messagebox.askyesno("保存测试结果", msg):
            # 保存已测试的结果
            self._save_partial_test_results(session)
        else:
            self.logger.info("用户选择不保存测试结果")
        
        self.end_test_progress()
    
    def _save_partial_test_results(self, session: TestSession):
        """保存部分测试结果"""
        try:
            self.logger.info("开始保存已测试的结果...")
            
            # 更新HAR相关统计
            self.tester.apply_har_metadata(session)
            
            # 计算元数据
            session.calculate_metadata()
            
            # 保存JSON文件
            output_dir = self._get_output_directory()
            filename = self.tester.save_results(session, output_dir)
            self.logger.info(f"结果已保存到: {filename}")
            
            # 获取对比文件路径（如果有）
            reference_file = ""
            if self.current_test_config:
                reference_file = self.current_test_config.get('reference_file', '')
            
            # 处理黑名单对比（如果有配置对比文件）
            if reference_file:
                self.logger.info("开始对比黑名单...")
                self.compare_and_prompt_blacklist(session, reference_file)
            
            # 更新UI显示（不调用 end_test_progress，因为已经在 _handle_stopped_test 中调用了）
            # 提取所有测试结果
            all_results = []
            for url_results in session.results.values():
                all_results.extend(url_results)
            
            # 存储当前会话结果
            self.current_session_name = session.test_name
            self.current_session_results = all_results
            self.all_sessions[session.test_name] = all_results
            self.session_objects[session.test_name] = session
            
            # 更新UI
            self.update_folder_combo()
            self.display_current_session()
            self.refresh_blacklist_display()
            
        except Exception as e:
            self.logger.error(f"保存测试结果时发生错误: {e}", exc_info=True)
            messagebox.showerror("保存失败", f"保存测试结果时发生错误:\n{str(e)}")

    def _run_single_test(self, url: str, url_index: int, original_position: int, round_num: int) -> TestResult:
        """
        在单独线程中运行单个URL测试
        
        注意：每个线程需要创建自己的测试器实例和Playwright实例，因为Playwright的同步API不是线程安全的
        """
        task_id = f"{original_position}-{round_num}"
        
        # 为每个线程创建任务专用的日志适配器
        from utils.logger import create_task_logger
        task_logger = create_task_logger(self.logger, task_id, f"URL-{original_position}")
        
        # 为每个线程创建独立的测试器实例
        thread_tester = PageLoadTester(logger=task_logger, wait_for_network_idle=self.tester.wait_for_network_idle, config_manager=self.config_manager)
        thread_tester.timeout = self.tester.timeout
        if hasattr(self.tester, "stage_timeouts"):
            thread_tester.stage_timeouts = self.tester.stage_timeouts.copy()
        thread_tester.headless = self.tester.headless
        thread_tester.update_har_options(self.tester.har_options)
        thread_tester.blacklist_manager = self.tester.blacklist_manager
        
        # 设置会话目录（共享同一个会话）
        if self.tester.current_session_name:
            thread_tester.current_session_name = self.tester.current_session_name
            thread_tester.session_output_dir = self.tester.session_output_dir
            if self.tester.har_manager:
                thread_tester.har_manager = self.tester.har_manager
        
        try:
            # 执行测试
            result = thread_tester._measure_single_page_with_retry(
                url, url_index, original_position, round_num, max_retries=1
            )
            
            # 记录任务完成状态
            if hasattr(task_logger, 'log_task_complete'):
                status = result.status if hasattr(result, 'status') else "未知"
                duration = result.response_time if hasattr(result, 'response_time') and result.response_time > 0 else -1
                task_logger.log_task_complete(url, original_position, round_num, status, duration)
            
            return result
        except Exception as e:
            # 记录任务错误
            if hasattr(task_logger, 'log_task_error'):
                task_logger.log_task_error(url, original_position, round_num, str(e))
            else:
                task_logger.error(f"任务执行失败: {e}")
            raise

    def _handle_test_completion(self, test_session, test_name, filename, reference_file: str = ""):
        """在主线程中处理测试完成"""
        # 结束进度跟踪
        self.end_test_progress()

        # 当前精简模式下，不再在UI中展示详细测试结果列表，仅记录日志。
        self.logger.info(f"测试 '{test_name}' 完成，结果已写入 HAR 文件目录: {filename}")
        # 更新当前测试输出目录
        self.output_dir_var.set(str(filename))
        self.status_var.set(f"测试 '{test_name}' 已完成（仅保存HAR文件）")

    def _handle_test_error(self, error_message):
        """在主线程中处理测试错误"""
        # 结束进度跟踪
        self.end_test_progress()

        messagebox.showerror("测试错误", f"运行测试时发生错误: {error_message}")
        self.logger.error(f"测试异常: {error_message}")
        self.status_var.set("测试失败")

    # ==================== 其他现有方法 ====================

    def setup_detail_columns(self):
        """设置详细视图的列"""
        # 清空现有列
        for col in self.result_tree['columns']:
            self.result_tree.heading(col, text="")
            self.result_tree.column(col, width=0)

        # 设置详细视图列标题
        self.result_tree['columns'] = self.detail_columns

        self.result_tree.heading("test_round", text="轮次")
        self.result_tree.heading("original_position", text="位置",
                                 command=lambda: self.sort_detail_by_column("original_position"))
        self.result_tree.heading("url", text="URL")
        self.result_tree.heading("final_url", text="最终URL")
        self.result_tree.heading("ip", text="IP地址")
        self.result_tree.heading("test_mode", text="网络环境")
        self.result_tree.heading("status", text="状态码",
                                 command=lambda: self.sort_detail_by_column("status"))
        self.result_tree.heading("fcp", text="FCP(s)", command=lambda: self.sort_detail_by_column("fcp"))
        self.result_tree.heading("dom_ready", text="DOM准备(s)", command=lambda: self.sort_detail_by_column("dom_ready"))
        self.result_tree.heading("full_load", text="load(s)", command=lambda: self.sort_detail_by_column("full_load"))

        # 设置列宽
        self.result_tree.column("test_round", width=50)
        self.result_tree.column("original_position", width=50)
        self.result_tree.column("url", width=180)
        self.result_tree.column("final_url", width=180)
        self.result_tree.column("ip", width=120)
        self.result_tree.column("test_mode", width=80)
        self.result_tree.column("status", width=70)
        self.result_tree.column("fcp", width=80)
        self.result_tree.column("dom_ready", width=80)
        self.result_tree.column("full_load", width=80)

    def setup_average_columns(self):
        """设置平均视图的列"""
        # 清空现有列
        for col in self.result_tree['columns']:
            self.result_tree.heading(col, text="")
            self.result_tree.column(col, width=0)

        # 设置平均视图列标题
        self.result_tree['columns'] = self.average_columns
        self.result_tree.heading("url", text="URL")
        self.result_tree.heading("ip", text="IP地址")
        self.result_tree.heading("position_range", text="位置范围",
                                 command=lambda: self.sort_average_by_column("position_range"))
        self.result_tree.heading("test_count", text="测试次数",
                                 command=lambda: self.sort_average_by_column("test_count"))
        self.result_tree.heading("avg_response", text="平均响应",
                                 command=lambda: self.sort_average_by_column("avg_response"))
        self.result_tree.heading("avg_dom", text="平均DOM",
                                 command=lambda: self.sort_average_by_column("avg_dom"))
        self.result_tree.heading("avg_fcp", text="平均FCP",
                                 command=lambda: self.sort_average_by_column("avg_fcp"))
        self.result_tree.heading("avg_full_load", text="平均load",
                                 command=lambda: self.sort_average_by_column("avg_full_load"))
        self.result_tree.heading("success_rate", text="成功率%",
                                 command=lambda: self.sort_average_by_column("success_rate"))

        # 设置列宽及对齐
        self.result_tree.column("url", width=220, anchor=tk.W)
        self.result_tree.column("ip", width=120, anchor=tk.W)
        self.result_tree.column("position_range", width=90, anchor=tk.CENTER)
        self.result_tree.column("test_count", width=90, anchor=tk.CENTER)
        self.result_tree.column("avg_response", width=90, anchor=tk.CENTER)
        self.result_tree.column("avg_dom", width=90, anchor=tk.CENTER)
        self.result_tree.column("avg_fcp", width=90, anchor=tk.CENTER)
        self.result_tree.column("avg_full_load", width=90, anchor=tk.CENTER)
        self.result_tree.column("success_rate", width=90, anchor=tk.CENTER)

    def sort_detail_by_column(self, column: str):
        """对详细视图的列进行排序"""
        if self.current_view != "detail":
            return

        reverse = self.detail_sort_states.get(column, False)
        valid_items = []
        invalid_items = []

        for item in self.result_tree.get_children(""):
            value = self.result_tree.set(item, column)
            if column == "original_position":
                sort_key = self._parse_position_value(value)
            elif column == "status":
                sort_key = self._parse_status_sort_value(value)
            else:
                sort_key = self._parse_sortable_value(value)

            if sort_key is None:
                invalid_items.append(item)
            else:
                valid_items.append((sort_key, item))

        valid_items.sort(key=lambda x: x[0], reverse=reverse)
        new_order = [iid for _, iid in valid_items] + invalid_items

        for index, iid in enumerate(new_order):
            self.result_tree.move(iid, '', index)

        self.detail_sort_states[column] = not reverse

    def sort_average_by_column(self, column: str):
        """对平均视图的列进行排序"""
        if self.current_view != "average":
            return

        reverse = self.average_sort_states.get(column, False)
        valid_items = []
        invalid_items = []

        for item in self.result_tree.get_children(""):
            value = self.result_tree.set(item, column)
            if column == "position_range":
                sort_key = self._parse_position_range_value(value)
            elif column == "test_count":
                sort_key = self._parse_test_count_value(value)
            elif column == "success_rate":
                cleaned = value.replace("%", "") if isinstance(value, str) else value
                sort_key = self._parse_sortable_value(cleaned)
            else:
                sort_key = self._parse_sortable_value(value)

            if sort_key is None:
                invalid_items.append(item)
            else:
                valid_items.append((sort_key, item))

        valid_items.sort(key=lambda x: x[0], reverse=reverse)
        new_order = [iid for _, iid in valid_items] + invalid_items

        for index, iid in enumerate(new_order):
            self.result_tree.move(iid, '', index)

        self.average_sort_states[column] = not reverse

    @staticmethod
    def _parse_sortable_value(value: Optional[str]) -> Optional[float]:
        """将列值转换为可排序的浮点数，-1或无效值返回None"""
        if value in ("", "-", "N/A", None):
            return None
        try:
            if isinstance(value, str):
                stripped = value.strip()
                if stripped.endswith("%"):
                    stripped = stripped[:-1]
                value = stripped
            numeric_value = float(value)
        except (TypeError, ValueError):
            return None
        if numeric_value < 0:
            return None
        return numeric_value

    @staticmethod
    def _parse_position_value(value: Optional[str]) -> Optional[int]:
        if value in ("", "-", None):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_position_range_value(value: Optional[str]) -> Optional[tuple]:
        if value in ("", "-", None):
            return None
        try:
            if "-" in value:
                start_str, end_str = value.split("-", 1)
                start = int(start_str)
                end = int(end_str) if end_str else start
                return (start, end)
            else:
                pos = int(value)
                return (pos, pos)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_test_count_value(value: Optional[str]) -> Optional[tuple]:
        if value in ("", "-", None):
            return None
        if isinstance(value, str) and "/" in value:
            parts = value.split("/", 1)
            try:
                total = int(parts[0])
                success = int(parts[1])
                return (total, success)
            except (TypeError, ValueError):
                return None
        try:
            total = int(value)
            return (total, 0)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_status_sort_value(value: Optional[str]) -> Optional[tuple]:
        if value in ("", None):
            return None
        value_str = str(value)
        if value_str.isdigit():
            return (0, int(value_str))

        priority_map = {
            "Timeout": 1,
            "部分成功": 2,
            "重定向错误": 3,
            "DNS错误": 4,
            "Error": 5,
        }
        return (1, priority_map.get(value_str, value_str))

    def toggle_view(self):
        """切换视图模式"""
        if self.current_view == "detail":
            self.current_view = "average"
            self.view_toggle_btn.config(text="切换到详细视图")
            self.view_status_var.set("视图: 平均视图")
        else:
            self.current_view = "detail"
            self.view_toggle_btn.config(text="切换到平均视图")
            self.view_status_var.set("视图: 详细视图")

        # 刷新当前显示
        self.display_current_session()

    def update_session_combo(self):
        """
        更新会话选择下拉框（当前精简版UI中已不再显示会话下拉框，此方法仅保留以兼容旧逻辑）

        这里不再操作任何UI控件，只是确保内部的会话名称列表是最新的。
        """
        # 仍然构造一次会话名称列表，方便后续如果需要使用
        _ = list(self.all_sessions.keys())
    
    def update_folder_combo(self):
        """
        更新文件夹选择下拉框

        当前精简版UI中已移除了左侧的会话/结果选择区域，此方法仅保留以兼容旧逻辑，
        不再操作任何具体的UI控件。
        """
        # 仍然扫描一次 results 目录，便于后续如果需要在其他地方使用这些信息
        results_dir = get_app_base_dir() / "results"
        if not results_dir.exists():
            return

        folders = set()
        folders.add("results")
        for json_file in results_dir.rglob("*.json"):
            rel_path = json_file.relative_to(results_dir)
            folder_path = rel_path.parent
            if folder_path == Path('.'):
                folders.add("results")
            else:
                folders.add(f"results/{folder_path}")

        # 只保留计算出的文件夹列表，不再写回到任何 Combobox
        _ = sorted(list(folders))
    
    def on_folder_selected(self, event):
        """当选择文件夹时的处理"""
        selected_folder = self.folder_var.get()
        if not selected_folder:
            return
        
        results_dir = get_app_base_dir() / "results"
        if not results_dir.exists():
            return
        
        # 确定实际文件夹路径
        if selected_folder == "results":
            folder_path = results_dir
        else:
            # 去掉"results/"前缀
            rel_path = selected_folder.replace("results/", "")
            folder_path = results_dir / rel_path
        
        self.selected_folder_path = selected_folder
        
        # 查找该文件夹下的所有JSON文件
        json_files = []
        if folder_path.exists():
            # 只查找当前文件夹下的JSON文件（不包括子文件夹）
            for json_file in folder_path.glob("*.json"):
                json_files.append(json_file.name)
        
        # 更新文件下拉框
        self.file_combo['values'] = sorted(json_files)
        self.file_var.set("")  # 清空文件选择
        
        # 清空当前会话
        self.current_session_name = ""
        self.current_session_results = []
        self.display_current_session()
    
    def on_file_selected(self, event):
        """当选择文件时的处理"""
        selected_file = self.file_var.get()
        if not selected_file:
            return
        
        selected_folder = self.folder_var.get()
        if not selected_folder:
            return
        
        results_dir = get_app_base_dir() / "results"
        
        # 确定实际文件夹路径
        if selected_folder == "results":
            folder_path = results_dir
        else:
            # 去掉"results/"前缀
            rel_path = selected_folder.replace("results/", "")
            folder_path = results_dir / rel_path
        
        # 构建完整文件路径
        file_path = folder_path / selected_file
        
        if not file_path.exists():
            messagebox.showerror("错误", f"文件不存在: {file_path}")
            return
        
        try:
            # 加载文件
            test_session = self.tester.load_results(file_path)
            
            # 提取所有测试结果
            all_results = []
            for url_results in test_session.results.values():
                all_results.extend(url_results)
            
            # 更新当前会话
            self.current_session_name = test_session.test_name
            self.current_session_results = all_results
            
            # 存储到all_sessions和session_objects（用于兼容性）
            self.all_sessions[test_session.test_name] = all_results
            self.session_objects[test_session.test_name] = test_session
            
            # 显示结果
            self.display_current_session()
            
            # 更新会话信息显示
            self.update_session_info_with_metadata()
            
            self.logger.info(f"成功加载会话: {test_session.test_name}")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载文件失败: {str(e)}")
            self.logger.error(f"加载文件失败: {str(e)}", exc_info=True)

    def on_session_selected(self, event):
        """当选择不同会话时的处理（保留用于兼容性）"""
        selected_session = self.session_var.get()
        if selected_session and selected_session in self.all_sessions:
            self.current_session_name = selected_session
            self.current_session_results = self.all_sessions[selected_session]
            self.display_current_session()

            # 更新会话信息显示
            self.update_session_info_with_metadata()

    def update_session_info_with_metadata(self):
        """使用完整的会话元数据更新信息显示"""
        if not self.current_session_name or self.current_session_name not in self.session_objects:
            return

        test_session = self.session_objects[self.current_session_name]

        # 更新会话信息文本框
        self.session_info_text.config(state=tk.NORMAL)
        self.session_info_text.delete(1.0, tk.END)

        self.session_info_text.insert(tk.END, f"会话名称: {test_session.test_name}\n")
        self.session_info_text.insert(tk.END, f"测试时间: {test_session.timestamp}\n")
        self.session_info_text.insert(tk.END, f"URL总数: {test_session.total_urls}\n")
        self.session_info_text.insert(tk.END, f"测试轮次: {test_session.test_rounds}\n")
        if test_session.session_directory:
            self.session_info_text.insert(tk.END, f"结果目录: {test_session.session_directory}\n")

        # 显示元数据统计
        if test_session.metadata:
            metadata = test_session.metadata
            self.session_info_text.insert(tk.END, f"总请求数: {metadata.get('total_requests', 0)}\n")
            self.session_info_text.insert(tk.END, f"成功请求: {metadata.get('successful_requests', 0)}\n")
            self.session_info_text.insert(tk.END, f"失败请求: {metadata.get('failed_requests', 0)}\n")
            self.session_info_text.insert(tk.END, f"成功率: {metadata.get('success_rate', 0)}%\n")
            self.session_info_text.insert(tk.END, f"平均响应: {metadata.get('average_response_time', 0):.3f}s\n")
            self.session_info_text.insert(tk.END, f"唯一URL: {metadata.get('unique_urls', 0)}\n")
            self.session_info_text.insert(tk.END, f"唯一IP: {metadata.get('unique_ips', 0)}\n")
            if metadata.get('har_files_count') is not None:
                self.session_info_text.insert(tk.END, f"HAR文件数: {metadata.get('har_files_count', 0)}\n")
            if metadata.get('total_unique_domains') is not None:
                self.session_info_text.insert(tk.END, f"唯一域名数: {metadata.get('total_unique_domains', 0)}\n")
        else:
            if test_session.har_files_count:
                self.session_info_text.insert(tk.END, f"HAR文件数: {test_session.har_files_count}\n")
            if test_session.total_unique_domains:
                self.session_info_text.insert(tk.END, f"唯一域名数: {test_session.total_unique_domains}\n")

        self.session_info_text.config(state=tk.DISABLED)

    def display_current_session(self):
        """显示当前会话的结果"""
        if not self.current_session_results:
            # 清空显示
            for item in self.result_tree.get_children():
                self.result_tree.delete(item)

            self.current_session_var.set("当前会话: 无")
            self.stats_var.set("")
            return

        # 更新会话信息
        self.current_session_var.set(f"当前会话: {self.current_session_name}")

        # 清空现有数据
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)

        if self.current_view == "detail":
            self.display_detail_view()
        else:
            self.display_average_view()

        # 更新会话信息文本
        self.update_session_info_with_metadata()

    def display_detail_view(self):
        """显示详细视图"""
        # 设置详细视图列
        self.setup_detail_columns()

        # 计算统计信息
        success_count = sum(1 for r in self.current_session_results if hasattr(r, 'status') and r.status == "success")
        error_count = len(self.current_session_results) - success_count
        rounds = len(set(r.test_round for r in self.current_session_results if r.test_round is not None))

        self.stats_var.set(f"成功: {success_count} | 失败: {error_count} | 轮次: {rounds}")

        # 按轮次和顺序排序结果
        sorted_results = sorted(self.current_session_results, key=lambda x: (
            x.test_round if x.test_round is not None else float('inf'),
            x.original_position if x.original_position is not None else float('inf')
        ))

        # 在添加结果到树状视图的部分，确保安全格式化
        for result in sorted_results:
            # 安全格式化性能指标，处理-1值
            fcp_display = "-1" if result.fcp_time == -1 else f"{result.fcp_time:.3f}" if result.fcp_time is not None else "-"
            dom_display = "-1" if result.dom_ready_time == -1 else f"{result.dom_ready_time:.3f}" if result.dom_ready_time is not None else "-"
            full_load_display = "-1" if result.full_load_time == -1 else f"{result.full_load_time:.3f}" if result.full_load_time is not None else "-"

            # 获取最终URL，如果没有则使用原始URL
            final_url = result.final_url or result.url

            # 确定状态显示内容
            # 使用统一的状态码格式化函数
            status_display = self._get_status_code_display(result)
            
            # 根据状态设置标签颜色
            if result.status == "success":
                tags = ("success",)
            elif result.status == "blocked":
                tags = ("blocked",)  # 拦截状态使用特殊标签
            elif result.status == "timeout":
                tags = ("error",)
            elif result.status == "partial_success":
                tags = ("warning",)
            elif result.error_type == "redirect_error":
                tags = ("redirect",)
            elif result.error_type == "dns_error":
                tags = ("dns_error",)
            else:  # error 状态
                tags = ("error",)

            # 确定网络环境显示
            test_mode_display = "直连网络"
            if hasattr(result, 'test_mode'):
                if result.test_mode == "vpn":
                    vpn_name = getattr(result, 'vpn_name', 'VPN')
                    test_mode_display = f"VPN({vpn_name})"
                elif result.test_mode == "direct":
                    test_mode_display = "直连网络"
            
            self.result_tree.insert(
                "", tk.END,
                values=(
                    result.test_round,
                    result.original_position,
                    result.url,
                    final_url,
                    result.ip_address or "-",
                    test_mode_display,
                    status_display,
                    fcp_display,
                    dom_display,
                    full_load_display
                ),
                tags=tags
            )

        # 配置标签样式
        self.result_tree.tag_configure("success", foreground="green")
        self.result_tree.tag_configure("error", foreground="red")
        self.result_tree.tag_configure("redirect", foreground="orange")  # 重定向错误用橙色
        self.result_tree.tag_configure("dns_error", foreground="purple")  # DNS错误用紫色
        self.result_tree.tag_configure("warning", foreground="brown")
        self.result_tree.tag_configure("blocked", foreground="darkred", background="lightyellow")  # 拦截状态用深红色文字，浅黄色背景

    def display_average_view(self):
        """显示平均视图"""
        # 设置平均视图列
        self.setup_average_columns()

        # 计算平均结果
        if self.current_session_name in self.session_objects:
            test_session = self.session_objects[self.current_session_name]
            self.current_average_results = test_session.calculate_average_results()
        else:
            # 如果没有完整的会话对象，从结果列表计算
            self.current_average_results = self.calculate_average_from_results()

        # 显示统计信息
        total_groups = len(self.current_average_results)
        total_tests = sum(result.test_count for result in self.current_average_results)
        total_success = sum(result.success_count for result in self.current_average_results)
        overall_success_rate = (total_success / total_tests * 100) if total_tests > 0 else 0

        self.stats_var.set(f"分组数: {total_groups} | 总测试数: {total_tests} | 总成功率: {overall_success_rate:.2f}%")

        # 显示平均结果
        for result in self.current_average_results:
            test_count_display = f"{result.test_count}/{result.success_count}"
            self.result_tree.insert(
                "", tk.END,
                values=(
                    result.url,
                    result.ip_address,
                    result.position_range,
                    test_count_display,
                    f"{result.avg_response_time:.2f}",
                    f"{result.avg_dom_ready_time:.2f}",
                    f"{result.avg_fcp_time:.2f}",
                    f"{result.avg_full_load_time:.2f}",
                    f"{result.success_rate:.2f}%"
                ),
                tags=("average",)
            )

        # 配置标签样式
        self.result_tree.tag_configure("average", foreground="blue")

    def calculate_average_from_results(self) -> List[AverageResult]:
        """从结果列表计算平均结果（备用方法）"""
        from collections import defaultdict

        # 按 (url, ip_address) 分组
        grouped_results = defaultdict(list)

        for result in self.current_session_results:
            key = (result.url, result.ip_address or "未知IP")
            grouped_results[key].append(result)

        average_results = []

        for (url, ip_address), results in grouped_results.items():
            # 计算成功和总测试次数
            successful_results = [r for r in results if hasattr(r, 'status') and r.status == "success"]
            total_count = len(results)
            success_count = len(successful_results)

            # 计算成功率：成功次数 / 总测试次数 × 100%
            success_rate = (success_count / total_count * 100) if total_count > 0 else 0

            # 计算各项指标的平均值（只基于成功的结果）
            if successful_results:
                avg_response_time = sum(
                    r.response_time for r in successful_results if
                    r.response_time is not None and r.response_time != -1) / len(successful_results)
                # 处理可能为None的性能指标
                dom_times = [r.dom_ready_time for r in successful_results if
                             r.dom_ready_time is not None and r.dom_ready_time != -1]
                avg_dom_ready_time = sum(dom_times) / len(dom_times) if dom_times else 0.0

                full_load_times = [r.full_load_time for r in successful_results if
                                   r.full_load_time is not None and r.full_load_time != -1]
                avg_full_load_time = sum(full_load_times) / len(full_load_times) if full_load_times else 0.0

                fcp_times = [r.fcp_time for r in successful_results if r.fcp_time is not None and r.fcp_time != -1]
                avg_fcp_time = sum(fcp_times) / len(fcp_times) if fcp_times else 0.0
            else:
                avg_response_time = avg_dom_ready_time = avg_full_load_time = avg_fcp_time = 0.0

            # 收集原始位置信息用于排序和显示
            original_positions = list(set(r.original_position for r in results if r.original_position is not None))
            original_positions.sort()

            # 生成位置范围显示
            if original_positions:
                if len(original_positions) == 1:
                    position_range = str(original_positions[0])
                else:
                    position_range = f"{min(original_positions)}-{max(original_positions)}"
            else:
                position_range = "-"

            average_results.append(AverageResult(
                url=url,
                ip_address=ip_address,
                avg_response_time=round(avg_response_time, 2),
                avg_dom_ready_time=round(avg_dom_ready_time, 2),
                avg_full_load_time=round(avg_full_load_time, 2),
                avg_fcp_time=round(avg_fcp_time, 2),
                test_count=total_count,
                success_count=success_count,  # 新增成功次数
                success_rate=round(success_rate, 2),
                original_positions=original_positions,
                position_range=position_range  # 新增位置范围
            ))

        # 按URL和原始位置排序
        average_results.sort(key=lambda x: (min(x.original_positions) if x.original_positions else float('inf'), x.url))

        return average_results

    def show_result_details(self, event):
        """显示选中结果的详细信息"""
        selected_item = self.result_tree.selection()
        if not selected_item:
            return

        self.detail_text.config(state=tk.NORMAL)
        self.detail_text.delete(1.0, tk.END)

        if self.current_view == "detail":
            self.show_detail_result_details(selected_item)
        else:
            self.show_average_result_details(selected_item)

        self.detail_text.config(state=tk.DISABLED)

    def show_detail_result_details(self, selected_item):
        """显示详细视图的选中结果详情"""
        # 获取选中的行数据
        item_values = self.result_tree.item(selected_item[0])["values"]
        selected_url = item_values[2]  # URL在第3列
        selected_round = item_values[0]  # 轮次在第1列
        selected_position = item_values[1]  # 位置在第2列
        selected_test_mode = item_values[5]  # 网络环境在第6列（test_mode）

        # 从网络环境显示文本中提取test_mode和vpn_name
        # 格式可能是: "VPN(名称)" 或 "直连网络"
        expected_test_mode = None
        expected_vpn_name = None
        
        if selected_test_mode:
            if selected_test_mode.startswith("VPN("):
                # 提取VPN名称，格式: "VPN(名称)"
                expected_test_mode = "vpn"
                vpn_name_part = selected_test_mode[4:-1]  # 去掉 "VPN(" 和 ")"
                expected_vpn_name = vpn_name_part
            elif selected_test_mode == "直连网络":
                expected_test_mode = "direct"

        # 查找完整结果（需要匹配URL、轮次、位置和网络环境）
        full_result = None
        for result in self.current_session_results:
            # 基本匹配条件：URL、轮次、位置必须匹配
            if not (result.url == selected_url and
                    result.test_round == selected_round and
                    result.original_position == selected_position):
                continue
            
            # 如果指定了网络环境，需要匹配网络环境
            if expected_test_mode:
                result_test_mode = getattr(result, 'test_mode', None)
                
                if expected_test_mode == "vpn":
                    # VPN模式：需要匹配test_mode和vpn_name
                    if result_test_mode == "vpn":
                        result_vpn_name = getattr(result, 'vpn_name', None)
                        if expected_vpn_name:
                            # 如果指定了VPN名称，必须精确匹配
                            if result_vpn_name == expected_vpn_name:
                                full_result = result
                                break
                        else:
                            # 如果没有指定VPN名称，匹配任意VPN
                            full_result = result
                            break
                elif expected_test_mode == "direct":
                    # 直连模式：需要匹配test_mode
                    if result_test_mode == "direct":
                        full_result = result
                        break
            else:
                # 如果没有网络环境信息，使用第一个匹配的结果（向后兼容）
                if not full_result:
                    full_result = result

        if not full_result:
            return

        # 显示详情
        self.detail_text.insert(tk.END, f"URL: {full_result.url}\n")
        self.detail_text.insert(tk.END, f"最终URL: {full_result.final_url or full_result.url}\n")  # 新增最终URL显示
        self.detail_text.insert(tk.END, f"IP地址: {full_result.ip_address or '未知'}\n")
        self.detail_text.insert(tk.END, f"轮次: {full_result.test_round}\n")
        self.detail_text.insert(tk.END, f"位置: {full_result.original_position}\n")
        
        # 显示网络环境信息
        result_test_mode = getattr(full_result, 'test_mode', None)
        if result_test_mode == "vpn":
            vpn_name = getattr(full_result, 'vpn_name', 'VPN')
            self.detail_text.insert(tk.END, f"网络环境: VPN ({vpn_name})\n")
        elif result_test_mode == "direct":
            self.detail_text.insert(tk.END, f"网络环境: 直连网络\n")
        else:
            self.detail_text.insert(tk.END, f"网络环境: 普通测试\n")
        
        self.detail_text.insert(tk.END, f"测试状态: {full_result.status}\n")
        if full_result.domain_count:
            self.detail_text.insert(tk.END, f"唯一域名数: {full_result.domain_count}\n")
        if full_result.har_file_path:
            self.detail_text.insert(tk.END, f"HAR文件: {full_result.har_file_path}\n")
        if full_result.hostname_file_path:
            self.detail_text.insert(tk.END, f"Hostname文件: {full_result.hostname_file_path}\n")

        # 使用统一的状态码格式化函数
        status_code_display = self._get_status_code_display(full_result)
        self.detail_text.insert(tk.END, f"HTTP状态码: {status_code_display}\n")

        # 安全地显示响应时间
        response_time_display = "-1" if full_result.response_time == -1 else f"{full_result.response_time:.3f}" if full_result.response_time is not None else "N/A"
        self.detail_text.insert(tk.END, f"响应时间: {response_time_display} s\n")

        self.detail_text.insert(tk.END, "-" * 50 + "\n")

        # 显示性能指标（包括部分成功的情况）
        if full_result.status in ["success", "partial_success"]:
            self.detail_text.insert(tk.END, "性能指标 (按发生顺序):\n")

            # 安全地显示性能指标，处理-1值
            fcp_display = "-1" if full_result.fcp_time == -1 else f"{full_result.fcp_time:.3f}" if full_result.fcp_time is not None else "N/A"
            dom_display = "-1" if full_result.dom_ready_time == -1 else f"{full_result.dom_ready_time:.3f}" if full_result.dom_ready_time is not None else "N/A"
            full_load_display = "-1" if full_result.full_load_time == -1 else f"{full_result.full_load_time:.3f}" if full_result.full_load_time is not None else "N/A"

            self.detail_text.insert(tk.END, f"1. FCP: {fcp_display} s\n")
            self.detail_text.insert(tk.END, f"2. DOM准备时间: {dom_display} s\n")
            self.detail_text.insert(tk.END, f"3. load: {full_load_display} s\n")
            self.detail_text.insert(tk.END, f"4. 响应时间: {response_time_display} s\n")
            
            # 如果是部分成功，说明情况
            if full_result.status == "partial_success":
                self.detail_text.insert(tk.END, "\n注意: 页面部分成功，load超时但已获取状态码和部分性能数据\n")
            
            self.detail_text.insert(tk.END, "-" * 50 + "\n")
            self.detail_text.insert(tk.END, "原始数据:\n")
            # 格式化JSON字符串，避免科学计数法
            json_str = json.dumps(full_result.to_dict(), indent=2, ensure_ascii=False)
            json_str = self._format_scientific_notation(json_str)
            self.detail_text.insert(tk.END, json_str)
        else:
            self.detail_text.insert(tk.END, f"错误类型: {full_result.error_type or '未知'}\n")
            self.detail_text.insert(tk.END, f"错误信息: {full_result.error_message or '未知错误'}\n")
            self.detail_text.insert(tk.END, f"响应时间: {response_time_display} s\n")

    def show_average_result_details(self, selected_item):
        """显示平均视图的选中结果详情"""
        # 获取选中的行数据
        item_values = self.result_tree.item(selected_item[0])["values"]
        selected_url = item_values[0]  # URL在第1列
        selected_ip = item_values[1]  # IP在第2列

        # 查找对应的平均结果
        selected_average_result = None
        for result in self.current_average_results:
            if result.url == selected_url and result.ip_address == selected_ip:
                selected_average_result = result
                break

        if not selected_average_result:
            return

        # 显示平均结果详情
        self.detail_text.insert(tk.END, f"位置范围: {selected_average_result.position_range}\n")
        self.detail_text.insert(tk.END, f"URL: {selected_average_result.url}\n")
        self.detail_text.insert(tk.END, f"IP地址: {selected_average_result.ip_address}\n")
        self.detail_text.insert(tk.END, f"测试次数: {selected_average_result.test_count}\n")
        self.detail_text.insert(tk.END, f"成功次数: {selected_average_result.success_count}\n")
        self.detail_text.insert(tk.END, f"成功率: {selected_average_result.success_rate}%\n")
        self.detail_text.insert(tk.END, "-" * 50 + "\n")
        self.detail_text.insert(tk.END, "性能指标平均值:\n")
        self.detail_text.insert(tk.END, f"平均响应时间: {selected_average_result.avg_response_time:.2f} s\n")
        self.detail_text.insert(tk.END, f"平均DOM准备时间: {selected_average_result.avg_dom_ready_time:.2f} s\n")
        self.detail_text.insert(tk.END, f"平均load: {selected_average_result.avg_full_load_time:.2f} s\n")
        self.detail_text.insert(tk.END, f"平均FCP: {selected_average_result.avg_fcp_time:.2f} s\n")
        self.detail_text.insert(tk.END, "-" * 50 + "\n")

        # 查找该分组下的所有原始记录
        self.detail_text.insert(tk.END, "原始测试记录:\n")
        original_records = []
        for result in self.current_session_results:
            if (result.url == selected_url and
                    (result.ip_address == selected_ip or (result.ip_address is None and selected_ip == "未知IP"))):
                original_records.append(result)

        # 按轮次和位置排序
        original_records.sort(key=lambda x: (
            x.test_round if x.test_round is not None else float('inf'),
            x.original_position if x.original_position is not None else float('inf')
        ))

        for record in original_records:
            status_display = "成功" if hasattr(record, 'status') and record.status == "success" else "失败"
            status_code_display = self._get_status_code_display(record)
            response_time_display = "-1" if record.response_time == -1 else f"{record.response_time:.3f}" if record.response_time is not None else "N/A"
            
            # 格式化性能指标
            dom_ready_display = f"{record.dom_ready_time:.3f}" if record.dom_ready_time is not None and record.dom_ready_time != -1 else "N/A"
            full_load_display = f"{record.full_load_time:.3f}" if record.full_load_time is not None and record.full_load_time != -1 else "N/A"
            fcp_display = f"{record.fcp_time:.3f}" if record.fcp_time is not None and record.fcp_time != -1 else "N/A"

            self.detail_text.insert(tk.END,
                                    f"轮次{record.test_round} 位置{record.original_position}: {status_display} "
                                    f"(状态码: {status_code_display}, 响应: {response_time_display}s, "
                                    f"DOM: {dom_ready_display}s, load: {full_load_display}s, FCP: {fcp_display}s)\n"
                                    )

    def export_results(self):
        """导出当前会话结果 - 优化版本：同时导出详细和平均视图"""
        if not self.current_session_results:
            messagebox.showinfo("无结果", "没有可导出的测试数据")
            return

        # 生成文件名
        # 使用ISO 8601基本格式：YYYYMMDDTHHMMSS
        timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
        filename = f"{self.current_session_name}_{timestamp}.xlsx"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=filename
        )
        if not file_path:
            return

        try:
            # 同时导出详细视图和平均视图
            self.export_both_views_to_excel(Path(file_path))
            messagebox.showinfo("导出成功", f"详细视图和平均视图已保存到: {file_path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出错误: {str(e)}")
            self.logger.error(f"导出Excel失败: {str(e)}", exc_info=True)

    def export_both_views_to_excel(self, file_path: Path):
        """将详细视图和平均视图导出到同一个Excel文件的不同工作表"""
        import pandas as pd
        from openpyxl import load_workbook

        # 确保输出目录存在
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # 创建Excel写入器
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:

            # 1. 导出详细视图
            detail_data = self._prepare_detail_data_for_export()
            if detail_data:
                detail_df = pd.DataFrame(detail_data)
                detail_df.to_excel(writer, sheet_name='详细视图', index=False)

            # 2. 导出平均视图
            average_data = self._prepare_average_data_for_export()
            if average_data:
                average_df = pd.DataFrame(average_data)
                average_df.to_excel(writer, sheet_name='平均视图', index=False)

        # 调整列宽
        self._adjust_excel_columns(file_path)

    def _prepare_detail_data_for_export(self) -> List[Dict]:
        """准备详细视图数据用于导出"""
        detail_data = []

        # 按轮次和顺序排序结果
        sorted_results = sorted(self.current_session_results, key=lambda x: (
            x.test_round if x.test_round is not None else float('inf'),
            x.original_position if x.original_position is not None else float('inf')
        ))

        for result in sorted_results:
            # 处理-1值的显示
            response_time_display = "-1" if result.response_time == -1 else f"{result.response_time:.3f}" if result.response_time is not None else "-"
            fcp_display = "-1" if result.fcp_time == -1 else f"{result.fcp_time:.3f}" if result.fcp_time is not None else "-"
            dom_display = "-1" if result.dom_ready_time == -1 else f"{result.dom_ready_time:.3f}" if result.dom_ready_time is not None else "-"
            full_load_display = "-1" if result.full_load_time == -1 else f"{result.full_load_time:.3f}" if result.full_load_time is not None else "-"

            # 获取最终URL
            final_url = result.final_url or result.url

            row_data = {
                '轮次': result.test_round,
                '位置': result.original_position,
                'URL': result.url,
                '最终URL': final_url,  # 确保包含最终URL
                'IP地址': result.ip_address or "-",
                '状态': self._get_status_display(result.status),
                '状态码': self._get_status_code_display(result),
                '响应时间(s)': response_time_display,
                'FCP(s)': fcp_display,
                'DOM准备(s)': dom_display,
                'load(s)': full_load_display,
                '唯一域名数': result.domain_count or 0,
                'HAR文件': result.har_file_path or "-",
                'Hostname文件': result.hostname_file_path or "-",
                '测试时间': result.timestamp
            }
            detail_data.append(row_data)

        return detail_data

    def _prepare_average_data_for_export(self) -> List[Dict]:
        """准备平均视图数据用于导出"""
        average_data = []

        # 计算平均结果
        if self.current_session_name in self.session_objects:
            test_session = self.session_objects[self.current_session_name]
            average_results = test_session.calculate_average_results()
        else:
            average_results = self.calculate_average_from_results()

        for result in average_results:
            test_count_display = f"{result.test_count}/{result.success_count}"
            row_data = {
                'URL': result.url,
                'IP地址': result.ip_address,
                '位置范围': result.position_range,
                '测试次数': test_count_display,
                '平均响应': f"{result.avg_response_time:.2f}",
                '平均DOM': f"{result.avg_dom_ready_time:.2f}",
                '平均FCP': f"{result.avg_fcp_time:.2f}",
                '平均load': f"{result.avg_full_load_time:.2f}",
                '成功率%': f"{result.success_rate:.2f}%"
            }
            average_data.append(row_data)

        return average_data

    def _get_status_display(self, status: str) -> str:
        """获取状态显示文本"""
        status_map = {
            "success": "成功",
            "error": "错误",
            "timeout": "超时",
            "blocked": "拦截",
            "partial_success": "部分成功"
        }
        return status_map.get(status, status)
    
    def _format_scientific_notation(self, json_str: str) -> str:
        """格式化JSON字符串中的科学计数法，避免显示科学计数法格式的小浮点数"""
        import re
        # 匹配科学计数法格式：数字e+/-数字（包括正负指数）
        pattern = r'\b\d+\.\d+e[+-]\d+\b'
        def replace_scientific(match):
            num = float(match.group(0))
            # 对于很小的浮点数（绝对值小于0.001），使用更多小数位避免科学计数法
            if abs(num) > 0 and abs(num) < 0.001:
                # 格式化为最多15位小数，去除尾随零
                formatted = f"{num:.15f}".rstrip('0').rstrip('.')
                return formatted if formatted != '' else '0'
            else:
                # 对于较大的浮点数，保持原格式
                return match.group(0)
        return re.sub(pattern, replace_scientific, json_str)

    def _get_status_code_display(self, result: TestResult) -> str:
        """获取状态码显示文本（统一格式化）"""
        # 如果URL被拦截，显示"拦截"
        if result.status == "blocked" or result.error_type == "blacklist_blocked":
            return "拦截"
        
        # 如果有实际状态码，优先显示状态码（即使状态是timeout，如果阶段1成功获取了状态码也要显示）
        if result.status_code is not None:
            # 如果是超时状态但有状态码，说明阶段1成功但后续阶段超时
            if result.status == "timeout" and result.error_type in ("dom_timeout", "load_timeout"):
                # 显示状态码，但可以添加说明（可选）
                return str(result.status_code)
            return str(result.status_code)
        
        # 根据状态显示对应的文本（去掉图标，纯文字显示）
        if result.status == "success":
            return "200"
        elif result.status == "timeout":
            # 根据错误类型显示更具体的超时信息
            if result.error_type == "connection_timeout":
                return "连接超时"
            elif result.error_type == "dom_timeout":
                return "DOM超时"
            elif result.error_type == "load_timeout":
                return "Load超时"
            else:
                return "超时"
        elif result.status == "partial_success":
            return "部分成功"
        elif result.status == "error":
            if result.error_type == "redirect_error":
                return "重定向错误"
            elif result.error_type == "dns_error":
                return "DNS错误"
            elif result.error_type == "http_error":
                return "HTTP错误"
            else:
                return "错误"
        else:
            return "N/A"

    def _adjust_excel_columns(self, file_path: Path):
        """调整Excel列宽"""
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        wb = load_workbook(file_path)

        # 设置列宽
        column_widths = {
            '详细视图': {
                'A': 8,  # 轮次
                'B': 8,  # 位置
                'C': 30,  # URL
                'D': 30,  # 最终URL
                'E': 15,  # IP地址
                'F': 8,  # 状态
                'G': 8,  # 状态码
                'H': 12,  # 响应时间
                'I': 12,  # 首次绘制
                'J': 12,  # DOM准备
                'K': 12,  # 完全加载
                'L': 12,  # 唯一域名数
                'M': 30,  # HAR文件
                'N': 30,  # Hostname文件
                'O': 20  # 测试时间
            },
            '平均视图': {
                'A': 40,  # URL
                'B': 15,  # IP地址
                'C': 12,  # 位置范围
                'D': 12,  # 测试次数
                'E': 15,  # 平均响应
                'F': 15,  # 平均DOM
                'G': 15,  # 平均FCP
                'H': 15,  # 平均load
                'I': 12  # 成功率
            }
        }

        for sheet_name, widths in column_widths.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for col_letter, width in widths.items():
                    ws.column_dimensions[col_letter].width = width

        wb.save(file_path)

    def open_results_dir(self):
        """打开结果存储目录"""
        results_dir = get_app_base_dir() / "results"
        if results_dir.exists() and results_dir.is_dir():
            try:
                import os
                os.startfile(results_dir)  # Windows打开目录
            except AttributeError:
                # 跨平台支持
                import subprocess
                import platform
                system = platform.system()
                if system == "Darwin":  # macOS
                    subprocess.call(["open", results_dir])
                elif system == "Linux":  # Linux
                    subprocess.call(["xdg-open", results_dir])
                else:  # 其他系统
                    messagebox.showinfo("打开目录", f"请手动打开目录: {results_dir}")
        else:
            messagebox.showinfo("无结果", "结果目录不存在")

    def load_previous_results(self, auto_load=False):
        """加载历史测试结果
        
        Args:
            auto_load: 如果为True，自动加载results目录下的所有文件（程序启动时使用）
                      如果为False，弹出对话框让用户选择文件夹和文件（手动加载时使用）
        """
        results_dir = get_app_base_dir() / "results"
        if not results_dir.exists():
            # 如果默认目录不存在，创建它
            results_dir.mkdir(parents=True, exist_ok=True)
            return

        if auto_load:
            # 自动加载模式：加载所有JSON文件（程序启动时使用）
            # 清空现有会话
            self.all_sessions.clear()
            self.session_objects = {}  # 存储完整的 TestSession 对象

            # 加载所有JSON结果文件（递归搜索所有子文件夹，排除对比文件）
            loaded_count = 0
            for file in results_dir.rglob("*.json"):
                try:
                    # 检查文件是否为空
                    if file.stat().st_size == 0:
                        self.logger.warning(f"跳过空文件: {file}")
                        continue

                    # 使用测试器的load_results方法加载完整的 TestSession
                    test_session = self.tester.load_results(file)

                    # 提取所有测试结果
                    all_results = []
                    for url_results in test_session.results.values():
                        all_results.extend(url_results)

                    # 存储会话结果和完整的会话对象
                    self.all_sessions[test_session.test_name] = all_results
                    self.session_objects[test_session.test_name] = test_session  # 存储完整会话对象
                    loaded_count += 1
                    self.logger.info(f"成功加载会话: {test_session.test_name}")

                except Exception as e:
                    self.logger.error(f"加载结果文件 {file} 失败: {str(e)}")
                    continue

            # 更新会话选择下拉框
            self.update_session_combo()

            # 更新文件夹选择下拉框
            self.update_folder_combo()

            self.logger.info(f"已加载 {loaded_count} 个测试会话")
        else:
            # 交互式加载模式：让用户选择文件夹和文件
            try:
                # 第一步：选择文件夹对话框
                folder_path = filedialog.askdirectory(
                    title="第一步：选择包含测试结果的文件夹",
                    initialdir=str(results_dir)
                )

                if not folder_path:
                    return  # 用户取消了选择

                folder = Path(folder_path)

                # 查找文件夹中的JSON文件（包括子文件夹）
                json_files = list(folder.rglob("*.json"))

                if not json_files:
                    messagebox.showinfo("提示", f"在文件夹 '{folder_path}' 中未找到JSON文件")
                    return

                # 第二步：选择文件对话框
                # 使用 askopenfilenames 让用户选择文件（支持多选）
                # 注意：在文件选择对话框中可以浏览子文件夹，然后选择文件
                messagebox.showinfo(
                    "提示",
                    f"请在文件选择对话框中浏览文件夹并选择要加载的测试文件。\n\n"
                    f"当前文件夹：{folder_path}\n"
                    f"您可以在对话框中进入子文件夹，然后选择JSON文件（支持多选）。"
                )
                selected_files = filedialog.askopenfilenames(
                    title="第二步：选择要加载的测试文件（可在对话框中浏览子文件夹，支持多选）",
                    initialdir=str(folder),
                    filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")]
                )

                if not selected_files:
                    return  # 用户取消了选择

                # 清空现有会话
                self.all_sessions.clear()
                self.session_objects = {}  # 存储完整的 TestSession 对象

                # 显示进度
                self.status_var.set(f"正在加载会话...")
                self.root.update_idletasks()

                # 加载用户选择的文件
                loaded_count = 0
                for file_path in selected_files:
                    file = Path(file_path)
                    try:
                        # 检查文件是否为空
                        if file.stat().st_size == 0:
                            self.logger.warning(f"跳过空文件: {file}")
                            continue

                        # 使用测试器的load_results方法加载完整的 TestSession
                        test_session = self.tester.load_results(file)

                        # 检查是否已存在同名会话
                        session_name = test_session.test_name
                        if session_name in self.session_objects:
                            # 询问是否覆盖
                            if not messagebox.askyesno("会话已存在",
                                                       f"会话 '{session_name}' 已存在，是否覆盖？"):
                                continue

                        # 提取所有测试结果
                        all_results = []
                        for url_results in test_session.results.values():
                            all_results.extend(url_results)

                        # 存储会话结果和完整的会话对象
                        self.all_sessions[session_name] = all_results
                        self.session_objects[session_name] = test_session  # 存储完整会话对象
                        loaded_count += 1
                        self.logger.info(f"成功加载会话: {session_name}")

                    except Exception as e:
                        self.logger.error(f"加载结果文件 {file} 失败: {str(e)}")
                        continue

                # 更新文件夹选择下拉框
                self.update_folder_combo()

                if loaded_count > 0:
                    # 尝试选择最新加载的文件（如果可能）
                    # 注意：由于现在是文件夹+文件选择模式，这里不再自动选择
                    pass

                    self.status_var.set(f"成功加载 {loaded_count} 个会话")
                    messagebox.showinfo("完成", f"成功加载 {loaded_count} 个会话")
                else:
                    self.status_var.set("未加载任何会话")
                    messagebox.showinfo("提示", "未加载任何会话")

            except Exception as e:
                messagebox.showerror("错误", f"加载会话失败: {str(e)}")
                self.logger.error(f"加载会话失败: {str(e)}", exc_info=True)
                self.status_var.set("加载会话失败")

    def load_external_results(self):
        """加载外部文件夹中的测试结果"""
        try:
            # 第一步：选择文件夹对话框
            folder_path = filedialog.askdirectory(
                title="第一步：选择包含测试结果的文件夹"
            )

            if not folder_path:
                return  # 用户取消了选择

            folder = Path(folder_path)

            # 查找文件夹中的JSON文件（包括子文件夹）
            json_files = list(folder.rglob("*.json"))

            if not json_files:
                messagebox.showinfo("提示", f"在文件夹 '{folder_path}' 中未找到JSON文件")
                return

            # 第二步：选择文件对话框
            # 将文件路径转换为字符串列表，供文件选择对话框使用
            initial_dir = str(folder)
            
            # 使用 askopenfilenames 让用户选择文件（支持多选）
            selected_files = filedialog.askopenfilenames(
                title="第二步：选择要加载的测试文件（可多选）",
                initialdir=initial_dir,
                filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")]
            )

            if not selected_files:
                return  # 用户取消了选择

            # 显示进度
            self.status_var.set(f"正在加载外部会话...")
            self.root.update_idletasks()

            # 加载用户选择的文件
            loaded_count = 0
            for file_path in selected_files:
                file = Path(file_path)
                try:
                    # 检查文件是否为空
                    if file.stat().st_size == 0:
                        self.logger.warning(f"跳过空文件: {file}")
                        continue

                    # 使用测试器的load_results方法加载完整的 TestSession
                    test_session = self.tester.load_results(file)

                    # 检查是否已存在同名会话
                    session_name = test_session.test_name
                    if session_name in self.session_objects:
                        # 询问是否覆盖
                        if not messagebox.askyesno("会话已存在",
                                                   f"会话 '{session_name}' 已存在，是否覆盖？"):
                            continue

                    # 提取所有测试结果
                    all_results = []
                    for url_results in test_session.results.values():
                        all_results.extend(url_results)

                    # 存储会话结果和完整的会话对象
                    self.all_sessions[session_name] = all_results
                    self.session_objects[session_name] = test_session
                    loaded_count += 1
                    self.logger.info(f"成功加载外部会话: {session_name}")

                except Exception as e:
                    self.logger.error(f"加载结果文件 {file} 失败: {str(e)}")
                    continue

            # 更新文件夹选择下拉框
            self.update_folder_combo()

            if loaded_count > 0:
                self.status_var.set(f"成功加载 {loaded_count} 个外部会话")
                messagebox.showinfo("完成", f"成功加载 {loaded_count} 个外部会话\n\n请在主界面选择文件夹和文件来查看结果。")
            else:
                self.status_var.set("未加载任何外部会话")
                messagebox.showinfo("提示", "未加载任何外部会话")

        except Exception as e:
            messagebox.showerror("错误", f"加载外部会话失败: {str(e)}")
            self.logger.error(f"加载外部会话失败: {str(e)}", exc_info=True)
            self.status_var.set("加载外部会话失败")

    # ==================== 黑名单相关方法 ====================

    def refresh_blacklist_display(self):
        """刷新黑名单显示"""
        self.all_blocked_domains = self.blacklist_manager.get_blocked_domains()
        self.blacklist_count_var.set(str(len(self.all_blocked_domains)))
        self._filter_blacklist_display()

    def _filter_blacklist_display(self):
        """根据输入框内容过滤黑名单显示"""
        search_text = self.blacklist_entry.get().strip().lower()
        
        self.blacklist_listbox.delete(0, tk.END)
        
        if search_text:
            # 搜索模式：显示匹配的域名
            filtered = [d for d in self.all_blocked_domains if search_text in d.lower()]
            for domain in filtered:
                self.blacklist_listbox.insert(tk.END, domain)
        else:
            # 显示所有域名
            for domain in self.all_blocked_domains:
                self.blacklist_listbox.insert(tk.END, domain)

    def _on_blacklist_entry_change(self, event=None):
        """输入框内容变化时触发搜索"""
        self._filter_blacklist_display()

    def _on_blacklist_entry_return(self, event=None):
        """回车键处理：如果输入的是新域名则添加，否则执行搜索"""
        text = self.blacklist_entry.get().strip()
        if not text:
            return
        
        # 检查是否已存在
        if text.lower() in [d.lower() for d in self.all_blocked_domains]:
            # 已存在，只是搜索，不添加
            return
        
        # 不存在，尝试添加
        self.add_domain_to_blacklist()

    def _select_all_domains(self, event=None):
        """全选所有域名"""
        self.blacklist_listbox.selection_set(0, tk.END)
        return "break"  # 阻止默认行为

    def _copy_selected_domains(self, event=None):
        """复制选中的域名到剪贴板"""
        selected_indices = self.blacklist_listbox.curselection()
        if not selected_indices:
            return "break"
        
        selected_domains = [self.blacklist_listbox.get(i) for i in selected_indices]
        text_to_copy = "\n".join(selected_domains)
        
        self.root.clipboard_clear()
        self.root.clipboard_append(text_to_copy)
        self.logger.info(f"已复制 {len(selected_domains)} 个域名到剪贴板")
        return "break"

    def add_domain_to_blacklist(self):
        """手动添加域名到黑名单"""
        domain = self.blacklist_entry.get().strip()
        if not domain:
            messagebox.showwarning("输入错误", "请输入要添加的域名")
            return
        
        # 规范化域名
        clean_domain = BlacklistManager._normalize_domain(domain)
        if not clean_domain:
            messagebox.showwarning("输入错误", "无效的域名格式")
            return
        
        # 检查是否已存在
        if clean_domain in self.blacklist_manager.blocked_domains:
            messagebox.showinfo("提示", f"域名 '{clean_domain}' 已在黑名单中")
            return
        
        # 添加到黑名单
        self.blacklist_manager.add_domains([clean_domain])
        # 更新完整列表
        self.all_blocked_domains = self.blacklist_manager.get_blocked_domains()
        self.blacklist_count_var.set(str(len(self.all_blocked_domains)))
        # 刷新显示（保持当前搜索状态）
        self._filter_blacklist_display()
        # 清空输入框
        self.blacklist_entry.delete(0, tk.END)
        self.logger.info(f"手动添加域名到黑名单: {clean_domain}")

    def remove_selected_domain(self):
        """删除选中的域名（支持多选）"""
        selected_indices = self.blacklist_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("提示", "请先选择要删除的域名")
            return
        
        selected_domains = [self.blacklist_listbox.get(i) for i in selected_indices]
        
        if len(selected_domains) == 1:
            msg = f"确定要从黑名单中删除 '{selected_domains[0]}' 吗？"
        else:
            msg = f"确定要从黑名单中删除选中的 {len(selected_domains)} 个域名吗？"
        
        if messagebox.askyesno("确认", msg):
            for domain in selected_domains:
                self.blacklist_manager.remove_domain(domain)
            # 更新完整列表
            self.all_blocked_domains = self.blacklist_manager.get_blocked_domains()
            self.blacklist_count_var.set(str(len(self.all_blocked_domains)))
            # 刷新显示（保持当前搜索状态）
            self._filter_blacklist_display()
            self.logger.info(f"已从黑名单删除 {len(selected_domains)} 个域名")

    def clear_blacklist(self):
        """清空黑名单"""
        if messagebox.askyesno("确认", "确定要清空所有黑名单域名吗？"):
            self.blacklist_manager.clear_blacklist()
            self.all_blocked_domains = []
            self.blacklist_count_var.set("0")
            self._filter_blacklist_display()
            self.logger.info("已清空黑名单")

    def compare_and_prompt_blacklist(self, test_session: TestSession, reference_file: str):
        """对比hostname文件并提示添加黑名单"""
        ref_path = Path(reference_file)
        
        # 检查对比文件是否存在
        if not ref_path.exists():
            self.logger.warning(f"对比文件不存在: {reference_file}")
            messagebox.showwarning("文件缺失", f"对比文件不存在，无法进行对比分析:\n{reference_file}")
            return

        try:
            # 加载参考文件中的域名
            reference_hostnames = BlacklistManager.load_reference_file(ref_path)
            if not reference_hostnames:
                self.logger.warning(f"对比文件为空或格式错误: {reference_file}")
                return

            # 收集测试中生成的所有hostname文件
            session_dir = Path(test_session.session_directory) if test_session.session_directory else get_app_base_dir() / "results" / self._sanitize_session_name(test_session.test_name)
            hostname_dir = session_dir / "hostname"
            
            if not hostname_dir.exists():
                self.logger.info("未找到hostname文件目录，跳过对比")
                return

            # 收集所有hostname文件中的域名（支持多种命名格式）
            test_hostnames = set()
            # 查找所有可能的hostname文件格式
            hostname_patterns = [
                "hostnames_*.TXT",
                "hostnames_*.txt",
                "*_hostnames.txt",
                "*_hostnames.TXT"
            ]
            hostname_files = []
            for pattern in hostname_patterns:
                hostname_files.extend(hostname_dir.glob(pattern))
            
            if not hostname_files:
                self.logger.info(f"未找到hostname文件（目录: {hostname_dir}），跳过对比")
                return
            
            self.logger.info(f"找到 {len(hostname_files)} 个hostname文件")
            for hostname_file in hostname_files:
                file_hostnames = BlacklistManager.load_hostnames_from_file(hostname_file)
                if file_hostnames:
                    test_hostnames.update(file_hostnames)
                    self.logger.debug(f"从 {hostname_file.name} 提取了 {len(file_hostnames)} 个域名")

            if not test_hostnames:
                self.logger.info("测试中未提取到hostname，跳过对比")
                return
            
            self.logger.info(f"共提取到 {len(test_hostnames)} 个唯一域名")

            # 对比找出匹配的基准URL（参考文件中的URL，使用匹配级别2：子域名匹配，与模糊对比.py保持一致）
            matched_base_urls = BlacklistManager.compare_hostnames(
                test_hostnames, 
                reference_hostnames, 
                match_level='2',
                return_base_urls=True  # 返回参考文件中的基准URL
            )
            
            if not matched_base_urls:
                self.logger.info("未发现匹配的域名")
                return

            # 过滤掉已经在黑名单中的域名
            new_domains = [d for d in matched_base_urls if d not in self.blacklist_manager.blocked_domains]
            
            if not new_domains:
                self.logger.info(f"找到 {len(matched_base_urls)} 个匹配的域名，但都已存在于黑名单中")
                messagebox.showinfo("黑名单对比", f"找到 {len(matched_base_urls)} 个匹配的域名，但都已存在于黑名单中，无需更新")
                return

            # 提示用户确认添加
            domain_list = "\n".join(new_domains[:20])  # 最多显示20个
            if len(new_domains) > 20:
                domain_list += f"\n... 还有 {len(new_domains) - 20} 个域名"
            
            msg = (
                f"发现 {len(new_domains)} 个匹配的基准URL（共匹配到 {len(matched_base_urls)} 个，"
                f"其中 {len(new_domains)} 个不在黑名单中）\n\n"
                f"匹配的基准URL（参考文件中的域名）:\n{domain_list}\n\n"
                f"是否添加到黑名单？"
            )
            
            if messagebox.askyesno("添加黑名单", msg):
                self.blacklist_manager.add_domains(new_domains)
                # 更新完整列表
                self.all_blocked_domains = self.blacklist_manager.get_blocked_domains()
                self.blacklist_count_var.set(str(len(self.all_blocked_domains)))
                # 刷新显示（保持当前搜索状态）
                self._filter_blacklist_display()
                self.logger.info(f"已添加 {len(new_domains)} 个域名到黑名单")
                messagebox.showinfo("完成", f"已成功添加 {len(new_domains)} 个域名到黑名单")
            else:
                self.logger.info("用户取消添加黑名单")

        except Exception as e:
            self.logger.error(f"对比hostname文件失败: {e}", exc_info=True)
            messagebox.showerror("错误", f"对比hostname文件时发生错误: {str(e)}")

    def open_blacklist_compare_dialog(self):
        """打开对比黑名单对话框"""
        try:
            # 创建对话框窗口
            dialog = tk.Toplevel(self.root)
            dialog.title("对比黑名单")
            dialog.geometry("600x550")
            dialog.resizable(True, True)
            dialog.minsize(550, 450)
            dialog.transient(self.root)
            dialog.grab_set()

            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
            y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
            dialog.geometry(f"+{x}+{y}")

            # 创建可滚动的主框架
            canvas = tk.Canvas(dialog)
            scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # 绑定鼠标滚轮
            def _on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

            main_frame = ttk.Frame(scrollable_frame, padding=20)
            main_frame.pack(fill=tk.BOTH, expand=True)

            # 说明文字
            info_label = ttk.Label(
                main_frame,
                text="选择测试数据源（文件夹或单个文件）和参考文件进行对比",
                font=("Arial", 9),
                foreground="gray"
            )
            info_label.pack(anchor=tk.W, pady=(0, 15))

            # 选择数据源类型
            source_type_frame = ttk.LabelFrame(main_frame, text="数据源类型", padding=10)
            source_type_frame.pack(fill=tk.X, pady=(0, 10))

            # 数据源选择区域（动态更新）
            source_frame = ttk.LabelFrame(main_frame, text="数据源", padding=10)
            source_frame.pack(fill=tk.X, pady=(0, 10))
            
            compare_source_var = tk.StringVar()
            compare_source_type_var = tk.StringVar(value='har')
            
            source_entry = ttk.Entry(source_frame, textvariable=compare_source_var, width=50)
            source_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

            def update_source_label():
                """更新数据源标签和选择按钮"""
                source_type = compare_source_type_var.get()
                if source_type == 'har':
                    source_frame.config(text="HAR文件（可选择文件夹或单个文件）")
                else:  # txt
                    source_frame.config(text="hostname.txt文件（可选择文件夹或单个文件）")

            def select_source():
                source_type = compare_source_type_var.get()
                # 先询问是选择文件夹还是单个文件
                choice_dialog = tk.Toplevel(dialog)
                choice_dialog.title("选择方式")
                choice_dialog.geometry("300x150")
                choice_dialog.transient(dialog)
                choice_dialog.grab_set()
                
                # 居中显示
                choice_dialog.update_idletasks()
                x = (choice_dialog.winfo_screenwidth() - choice_dialog.winfo_width()) // 2
                y = (choice_dialog.winfo_screenheight() - choice_dialog.winfo_height()) // 2
                choice_dialog.geometry(f"+{x}+{y}")
                
                choice_frame = ttk.Frame(choice_dialog, padding=20)
                choice_frame.pack(fill=tk.BOTH, expand=True)
                
                choice_result = {"value": None}
                
                def select_folder():
                    if source_type == 'har':
                        path = filedialog.askdirectory(
                            title="选择包含HAR文件的文件夹",
                            initialdir=get_app_base_dir() / "results"
                        )
                    else:  # txt
                        path = filedialog.askdirectory(
                            title="选择包含hostname.txt文件的文件夹",
                            initialdir=get_app_base_dir() / "results"
                        )
                    if path:
                        compare_source_var.set(path)
                    choice_dialog.destroy()
                
                def select_file():
                    if source_type == 'har':
                        path = filedialog.askopenfilename(
                            title="选择HAR文件",
                            filetypes=[("HAR文件", "*.har"), ("所有文件", "*.*")],
                            initialdir=get_app_base_dir() / "results"
                        )
                    else:  # txt
                        path = filedialog.askopenfilename(
                            title="选择hostname.txt文件",
                            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
                            initialdir=get_app_base_dir() / "results"
                        )
                    if path:
                        compare_source_var.set(path)
                    choice_dialog.destroy()
                
                ttk.Label(choice_frame, text="请选择数据源方式：", font=("Arial", 10)).pack(pady=(0, 15))
                
                btn_frame = ttk.Frame(choice_frame)
                btn_frame.pack()
                
                ttk.Button(btn_frame, text="选择文件夹", command=select_folder, width=15).pack(side=tk.LEFT, padx=5)
                ttk.Button(btn_frame, text="选择单个文件", command=select_file, width=15).pack(side=tk.LEFT, padx=5)
                ttk.Button(btn_frame, text="取消", command=choice_dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)

            ttk.Button(source_frame, text="选择", command=select_source).pack(side=tk.LEFT)
            ttk.Radiobutton(
                source_type_frame,
                text="HAR文件（自动提取hostname）",
                variable=compare_source_type_var,
                value='har',
                command=update_source_label
            ).pack(anchor=tk.W, pady=2)
            ttk.Radiobutton(
                source_type_frame,
                text="hostname.txt文件",
                variable=compare_source_type_var,
                value='txt',
                command=update_source_label
            ).pack(anchor=tk.W, pady=2)
            
            # 初始化UI
            update_source_label()

            # 选择参考文件
            ref_file_frame = ttk.LabelFrame(main_frame, text="参考文件（基准文件）", padding=10)
            ref_file_frame.pack(fill=tk.X, pady=(0, 10))

            compare_ref_file_var = tk.StringVar()
            ref_entry = ttk.Entry(ref_file_frame, textvariable=compare_ref_file_var, width=50)
            ref_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

            def select_ref_file():
                file = filedialog.askopenfilename(
                    title="选择参考文件",
                    filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
                    initialdir=get_app_base_dir()
                )
                if file:
                    compare_ref_file_var.set(file)

            ttk.Button(ref_file_frame, text="选择文件", command=select_ref_file).pack(side=tk.LEFT)

            # 匹配级别选择
            match_level_frame = ttk.LabelFrame(main_frame, text="匹配级别", padding=10)
            match_level_frame.pack(fill=tk.X, pady=(0, 15))

            compare_match_level_var = tk.StringVar(value='2')
            ttk.Radiobutton(
                match_level_frame,
                text="级别1: 精确匹配",
                variable=compare_match_level_var,
                value='1'
            ).pack(anchor=tk.W, pady=2)
            ttk.Radiobutton(
                match_level_frame,
                text="级别2: 子域名匹配（推荐）",
                variable=compare_match_level_var,
                value='2'
            ).pack(anchor=tk.W, pady=2)
            ttk.Radiobutton(
                match_level_frame,
                text="级别3: 二级域名匹配",
                variable=compare_match_level_var,
                value='3'
            ).pack(anchor=tk.W, pady=2)

            # 按钮区域
            btn_frame = ttk.Frame(main_frame)
            btn_frame.pack(fill=tk.X, pady=(10, 0))

            def on_ok():
                source_path = compare_source_var.get().strip()
                source_type = compare_source_type_var.get()
                ref_file = compare_ref_file_var.get().strip()
                match_level = compare_match_level_var.get()

                if not source_path:
                    source_type_name = "HAR文件" if source_type == 'har' else "hostname.txt文件"
                    messagebox.showwarning("输入错误", f"请选择{source_type_name}或文件夹")
                    return

                if not ref_file:
                    messagebox.showwarning("输入错误", "请选择参考文件")
                    return

                if not Path(source_path).exists():
                    messagebox.showerror("错误", f"路径不存在: {source_path}")
                    return

                if not Path(ref_file).exists():
                    messagebox.showerror("错误", f"参考文件不存在: {ref_file}")
                    return

                dialog.destroy()
                # 执行对比
                # 根据用户选择的数据源类型和实际路径判断
                source_path_obj = Path(source_path)
                if source_type == 'har':
                    # HAR文件类型，保持为'har'
                    actual_source_type = 'har'
                elif source_type == 'txt':
                    # txt文件类型，保持为'hostname'
                    actual_source_type = 'hostname'
                else:
                    # 兼容旧逻辑：如果是文件夹，使用'folder'类型
                    if source_path_obj.is_dir():
                        actual_source_type = 'folder'
                    else:
                        # 根据文件扩展名判断
                        if source_path_obj.suffix.lower() == '.har':
                            actual_source_type = 'har'
                        else:
                            actual_source_type = 'hostname'
                
                self.manual_compare_blacklist(source_path, ref_file, match_level, actual_source_type)

            def on_cancel():
                dialog.destroy()

            ttk.Button(btn_frame, text="确定", command=on_ok, width=12).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(btn_frame, text="取消", command=on_cancel, width=12).pack(side=tk.RIGHT)

        except Exception as e:
            self.logger.error(f"打开对比黑名单对话框失败: {e}", exc_info=True)
            messagebox.showerror("错误", f"打开对话框时发生错误: {str(e)}")

    def manual_compare_blacklist(self, source_path: str, reference_file: str, match_level: str = '2', source_type: str = 'folder'):
        """
        手动对比黑名单（不依赖TestSession）
        
        Args:
            source_path: 数据源路径（文件夹、HAR文件或hostname.txt文件）
            reference_file: 参考文件路径
            match_level: 匹配级别
            source_type: 数据源类型 ('folder', 'har', 'hostname')
        """
        ref_path = Path(reference_file)
        source_path_obj = Path(source_path)

        # 检查对比文件是否存在
        if not ref_path.exists():
            self.logger.warning(f"对比文件不存在: {reference_file}")
            messagebox.showwarning("文件缺失", f"对比文件不存在，无法进行对比分析:\n{reference_file}")
            return

        try:
            # 加载参考文件中的域名
            reference_hostnames = BlacklistManager.load_reference_file(ref_path)
            if not reference_hostnames:
                self.logger.warning(f"对比文件为空或格式错误: {reference_file}")
                messagebox.showwarning("文件错误", "参考文件为空或格式错误")
                return

            self.logger.info(f"已加载参考文件: {len(reference_hostnames)} 个域名")

            # 根据数据源类型提取hostname
            test_hostnames = set()
            
            # 判断是文件夹还是文件
            is_folder = source_path_obj.is_dir()
            
            if source_type == 'har':
                # 处理HAR文件（文件夹或单个文件）
                from utils.har_parser import HARParser
                
                if is_folder:
                    # 递归处理文件夹及其子文件夹中的所有HAR文件
                    har_files = list(source_path_obj.rglob("*.har"))
                    if not har_files:
                        messagebox.showwarning("文件缺失", f"在以下文件夹及其子文件夹中未找到HAR文件:\n{source_path}")
                        return
                    
                    self.logger.info(f"找到 {len(har_files)} 个HAR文件（包括子文件夹）")
                    for har_file in har_files:
                        hostnames = HARParser.extract_hostnames(har_file)
                        for hostname in hostnames:
                            normalized = BlacklistManager._normalize_domain(hostname)
                            if normalized:
                                test_hostnames.add(normalized)
                        self.logger.debug(f"从 {har_file.relative_to(source_path_obj)} 提取了 {len(hostnames)} 个hostname")
                    
                    self.logger.info(f"从 {len(har_files)} 个HAR文件（包括子文件夹）提取到 {len(test_hostnames)} 个唯一hostname（规范化后）")
                else:
                    # 处理单个HAR文件
                    if not source_path_obj.exists() or not source_path_obj.suffix.lower() == '.har':
                        messagebox.showerror("文件错误", f"HAR文件不存在或格式错误:\n{source_path}")
                        return
                    
                    hostnames = HARParser.extract_hostnames(source_path_obj)
                    if not hostnames:
                        messagebox.showwarning("数据缺失", f"未能从HAR文件中提取到hostname:\n{source_path}")
                        return
                    
                    # 规范化hostname（去除www前缀等）
                    for hostname in hostnames:
                        normalized = BlacklistManager._normalize_domain(hostname)
                        if normalized:
                            test_hostnames.add(normalized)
                    
                    self.logger.info(f"从HAR文件提取到 {len(hostnames)} 个hostname，规范化后 {len(test_hostnames)} 个")
                
            elif source_type == 'hostname':
                # 处理hostname.txt文件（文件夹或单个文件）
                if is_folder:
                    # 递归处理文件夹及其子文件夹中的所有txt文件
                    txt_files = list(source_path_obj.rglob("*.txt"))
                    if not txt_files:
                        messagebox.showwarning("文件缺失", f"在以下文件夹及其子文件夹中未找到txt文件:\n{source_path}")
                        return
                    
                    self.logger.info(f"找到 {len(txt_files)} 个txt文件（包括子文件夹）")
                    for txt_file in txt_files:
                        file_hostnames = BlacklistManager.load_hostnames_from_file(txt_file)
                        if file_hostnames:
                            test_hostnames.update(file_hostnames)
                            self.logger.debug(f"从 {txt_file.relative_to(source_path_obj)} 提取了 {len(file_hostnames)} 个域名")
                    
                    self.logger.info(f"从 {len(txt_files)} 个txt文件（包括子文件夹）提取到 {len(test_hostnames)} 个唯一域名")
                else:
                    # 处理单个hostname.txt文件
                    if not source_path_obj.exists() or not source_path_obj.suffix.lower() == '.txt':
                        messagebox.showerror("文件错误", f"hostname文件不存在或格式错误:\n{source_path}")
                        return
                    
                    file_hostnames = BlacklistManager.load_hostnames_from_file(source_path_obj)
                    if not file_hostnames:
                        messagebox.showwarning("数据缺失", f"未能从hostname文件中提取到域名:\n{source_path}")
                        return
                    
                    test_hostnames = file_hostnames
                    self.logger.info(f"从hostname文件提取到 {len(test_hostnames)} 个域名")
                
            else:  # source_type == 'folder'
                # 递归从文件夹及其子文件夹中查找hostname文件
                if not source_path_obj.exists() or not source_path_obj.is_dir():
                    messagebox.showerror("文件夹错误", f"文件夹不存在:\n{source_path}")
                    return
                
                # 递归查找所有hostname文件（支持多种命名格式）
                # 递归查找所有子文件夹中的文件
                hostname_patterns = [
                    "hostnames_*.TXT",
                    "hostnames_*.txt",
                    "*_hostnames.txt",
                    "*_hostnames.TXT",
                    "*.txt"  # 也支持所有txt文件（递归查找）
                ]
                hostname_files = []
                for pattern in hostname_patterns:
                    # 使用rglob递归查找所有子文件夹中的匹配文件
                    hostname_files.extend(source_path_obj.rglob(pattern))

                # 去重文件列表（即使有相同名称的文件，也会合并处理）
                hostname_files = list(set(hostname_files))

                if not hostname_files:
                    self.logger.warning(f"未找到hostname文件（目录: {source_path_obj}，包括子文件夹）")
                    messagebox.showwarning("文件缺失", f"在以下目录及其子文件夹中未找到hostname文件:\n{source_path_obj}\n\n请确保该目录包含hostname.txt文件")
                    return

                self.logger.info(f"找到 {len(hostname_files)} 个hostname文件（包括子文件夹，相同名称的文件会合并处理）")
                for hostname_file in hostname_files:
                    file_hostnames = BlacklistManager.load_hostnames_from_file(hostname_file)
                    if file_hostnames:
                        test_hostnames.update(file_hostnames)
                        self.logger.debug(f"从 {hostname_file.relative_to(source_path_obj)} 提取了 {len(file_hostnames)} 个域名")

            if not test_hostnames:
                self.logger.warning("未提取到hostname")
                messagebox.showwarning("数据缺失", "未能从hostname文件中提取到域名")
                return

            self.logger.info(f"共提取到 {len(test_hostnames)} 个唯一域名")

            # 对比找出匹配的基准URL（参考文件中的URL）
            matched_base_urls = BlacklistManager.compare_hostnames(
                test_hostnames, 
                reference_hostnames, 
                match_level=match_level,
                return_base_urls=True  # 返回参考文件中的基准URL
            )

            if not matched_base_urls:
                self.logger.info("未发现匹配的域名")
                messagebox.showinfo("对比结果", "未发现匹配的域名")
                return

            # 过滤掉已经在黑名单中的域名
            new_domains = [d for d in matched_base_urls if d not in self.blacklist_manager.blocked_domains]

            if not new_domains:
                self.logger.info("所有匹配的域名已在黑名单中")
                messagebox.showinfo("对比结果", "所有匹配的域名已在黑名单中")
                return

            # 显示匹配结果
            domain_list = "\n".join(new_domains[:20])  # 最多显示20个
            if len(new_domains) > 20:
                domain_list += f"\n... 还有 {len(new_domains) - 20} 个域名"

            msg = (
                f"发现 {len(new_domains)} 个匹配的基准URL（共匹配到 {len(matched_base_urls)} 个，"
                f"其中 {len(new_domains)} 个不在黑名单中）\n\n"
                f"匹配的基准URL（参考文件中的域名）:\n{domain_list}\n\n"
                f"是否添加到黑名单？"
            )

            if messagebox.askyesno("添加黑名单", msg):
                self.blacklist_manager.add_domains(new_domains)
                # 更新完整列表
                self.all_blocked_domains = self.blacklist_manager.get_blocked_domains()
                self.blacklist_count_var.set(str(len(self.all_blocked_domains)))
                # 刷新显示
                self._filter_blacklist_display()
                self.logger.info(f"已添加 {len(new_domains)} 个域名到黑名单")
                messagebox.showinfo("完成", f"已成功添加 {len(new_domains)} 个域名到黑名单")
            else:
                self.logger.info("用户取消添加黑名单")

        except Exception as e:
            self.logger.error(f"对比hostname文件失败: {e}", exc_info=True)
            messagebox.showerror("错误", f"对比hostname文件时发生错误: {str(e)}")

    def on_closing(self):
        """处理窗口关闭事件"""
        # 检查是否有定时任务
        has_scheduled_tasks = self.scheduler.has_enabled_tasks()
        
        if self.is_testing:
            # 如果正在测试，询问用户是否确认关闭
            if has_scheduled_tasks:
                response = messagebox.askyesno(
                    "确认关闭",
                    "当前有测试正在运行，并且有定时任务已设置。\n\n"
                    "关闭程序将停止当前测试，但定时任务配置会保留。\n\n"
                    "是否确认关闭程序？"
                )
            else:
                response = messagebox.askyesno(
                    "确认关闭",
                    "当前有测试正在运行，关闭程序将停止测试。\n\n"
                    "是否确认关闭程序？"
                )
            if response:
                # 用户确认关闭，停止测试
                self.test_stopped = True
                if self.tester and hasattr(self.tester, "request_stop"):
                    self.tester.request_stop()
                self.logger.info("用户确认关闭程序，停止测试")
                # 等待一下让测试停止
                import time
                time.sleep(0.5)
                # 如果测试会话存在，询问是否保存
                if self.current_test_session:
                    self._handle_stopped_test(self.current_test_session)
                # 停止调度器
                self.scheduler.stop()
                # 等待后台测试线程完成，避免强制退出导致Playwright异常
                self._wait_for_test_threads()
                self.root.destroy()
        elif has_scheduled_tasks:
            # 有定时任务，提示用户
            tasks = self.scheduler.get_tasks()
            enabled_tasks = [t for t in tasks if t['enabled']]
            
            weekday_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            task_info = []
            for task in enabled_tasks:
                days = [weekday_names[d] for d in task.get('weekdays', [])]
                times = ', '.join(task.get('times', []))
                task_info.append(f"  - {task['test_name']}: {', '.join(days)} {times}")
            
            response = messagebox.askyesno(
                "确认关闭",
                f"您有 {len(enabled_tasks)} 个定时任务已启用：\n\n" +
                "\n".join(task_info) + "\n\n" +
                "关闭程序后，定时任务将不会执行。\n\n"
                "是否确认关闭程序？"
            )
            if response:
                # 停止调度器
                self.scheduler.stop()
                self._wait_for_test_threads()
                self.root.destroy()
        else:
            # 没有测试运行，也没有定时任务，直接关闭
            self.scheduler.stop()
            self._wait_for_test_threads()
            self.root.destroy()
    
    @staticmethod
    def _sanitize_session_name(name: str) -> str:
        """清理会话名称（与tester中的方法保持一致）"""
        safe = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        return safe.replace(' ', '_') if safe else "unnamed_test"